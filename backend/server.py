"""FastAPI control server: all API endpoints, overlay/frame/metrics management, FFmpeg publishing."""

import os
import yaml
import uuid
import shutil
import threading
import time
import subprocess
import json
import logging
import asyncio
import math
import re
import io
import tempfile
from typing import Dict, List, Optional
from pathlib import Path
from collections import deque
import multiprocessing as mp
from datetime import datetime

import httpx
import requests as http_requests
import numpy as np
import cv2
import psutil

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from login import (
    login_user, add_user, delete_user, change_password, list_users,
    create_session, get_session, touch_session, delete_session, purge_expired_sessions,
)
from sam import (
    load_sam_model, sam_annotate_frame, sam_worker,
    sam_model_loaded, sam_results_lock, sam_results, sam_threads,
)
from log_utils import env_log_path, new_log_path
from report import generate_report, REPORTS_DIR
from helpers import get_video_duration, get_video_fps, JPEG_QUALITY
import forensics as forensics_events
import realtime_forensics as rt_forensics
import crowd_report
import realtime_crowd as rt_crowd
import magicboxhub_api

IRIS_LOCAL = os.environ.get("IRIS_LOCAL", "0") == "1"
DEFAULT_MEDIAMTX_HOST = "127.0.0.1" if IRIS_LOCAL else "mediamtx1.stagingbot.xyz"
MEDIAMTX_HOST = os.environ.get("IRIS_MEDIAMTX_HOST", DEFAULT_MEDIAMTX_HOST)
DEFAULT_MEDIAMTX_API = f"http://{MEDIAMTX_HOST}:9997"
MEDIAMTX_API = os.environ.get("IRIS_MEDIAMTX_API", DEFAULT_MEDIAMTX_API)


# ── Pydantic models ──

class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreateRequest(BaseModel):
    username: str
    password: str

class UserDeleteRequest(BaseModel):
    username: str

class PasswordChangeRequest(BaseModel):
    username: str
    new_password: str

class OverlayUpdate(BaseModel):
    heatmap: Optional[bool] = None
    heatmap_full: Optional[bool] = None
    heatmap_trails: Optional[bool] = None
    trails: Optional[bool] = None
    bboxes: Optional[bool] = None
    confidence: Optional[float] = None

class SourceIndexRequest(BaseModel):
    index: int

class SourceStartRequest(BaseModel):
    index: int
    mode: Optional[str] = None

class UploadInitRequest(BaseModel):
    filename: str
    size: int
    mode: Optional[str] = None
    chunk_size: Optional[int] = None

class UploadCompleteRequest(BaseModel):
    upload_id: str

class ActiveSourcesRequest(BaseModel):
    indexes: List[int]

class ConfidenceUpdate(BaseModel):
    confidence: float

class SamStartRequest(BaseModel):
    source: str
    prompt: str
    confidence: float = 0.7
    show_boxes: bool = True
    show_masks: bool = True

class SamStopRequest(BaseModel):
    source: str

class SamUpdateRequest(BaseModel):
    source: str
    confidence: Optional[float] = None
    show_boxes: Optional[bool] = None
    show_masks: Optional[bool] = None

class FrontendLogEntry(BaseModel):
    level: str
    message: str
    context: Optional[dict] = None
    ts: Optional[str] = None


# ── Config paths ──

RTSP_CONFIG_PATH = Path("config/rtsp_links.yml")
DATA_DIR = Path("data")
UPLOAD_DIR = DATA_DIR / "uploads" / "recordings"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
FORENSICS_UPLOAD_DIR = DATA_DIR / "uploads" / "forensics"
FORENSICS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALERTS_DIR = DATA_DIR / "alerts"
ALERTS_DIR.mkdir(parents=True, exist_ok=True)


# ── Thread-safe state ──

alerts_lock = threading.Lock()
alerts: deque = deque(maxlen=50)
alert_cooldowns: Dict[str, float] = {}
ALERT_COOLDOWN_SECONDS = 15
ALERT_TRIGGER_THRESHOLD = 70

overlay_lock = threading.Lock()
overlays: Dict[str, Dict[str, bool]] = {}

metrics_lock = threading.Lock()
metrics: Dict[str, dict] = {}

frame_lock = threading.Condition()
frame_buffer: Dict[str, bytes] = {}
frame_sequences: Dict[str, int] = {}
frame_bgr_buffer: Dict[str, np.ndarray] = {}

raw_frame_lock = threading.Condition()
raw_frame_buffer: Dict[str, bytes] = {}
raw_frame_sequences: Dict[str, int] = {}
raw_frame_bgr_buffer: Dict[str, np.ndarray] = {}

jobs_lock = threading.Lock()
jobs: Dict[str, dict] = {}

source_lock = threading.Lock()
running_sources: Dict[int, dict] = {}

# FFmpeg RTSP publishers (processed stream -> MediaMTX)
ffmpeg_publishers: Dict[str, subprocess.Popen] = {}
ffmpeg_log_files: Dict[str, object] = {}
ffmpeg_writer_threads: Dict[str, threading.Thread] = {}
ffmpeg_stop_events: Dict[str, threading.Event] = {}
ffmpeg_next_start: Dict[str, float] = {}
ffmpeg_frame_sequences: Dict[str, int] = {}
ffmpeg_state_lock = threading.Lock()
ffmpeg_starting: set = set()
raw_ffmpeg_publishers: Dict[str, subprocess.Popen] = {}
raw_ffmpeg_log_files: Dict[str, object] = {}
raw_ffmpeg_writer_threads: Dict[str, threading.Thread] = {}
raw_ffmpeg_stop_events: Dict[str, threading.Event] = {}
MEDIAMTX_RTSP_PORT = int(os.environ.get("IRIS_MEDIAMTX_RTSP_PORT", "8554"))
# Always publish to the local MediaMTX instance (same machine).
# The remote hostname is only for frontend WebRTC consumption via Cloudflare.
MEDIAMTX_PUBLISH_BASE = os.environ.get(
    "IRIS_MEDIAMTX_PUBLISH_BASE",
    f"rtsp://127.0.0.1:{MEDIAMTX_RTSP_PORT}",
)
PROCESSED_FPS = int(os.environ.get("IRIS_PROCESSED_FPS", "30"))
KEYFRAME_INTERVAL_SEC = int(os.environ.get("IRIS_KEYFRAME_INTERVAL_SEC", "1"))
PROCESSED_BITRATE = os.environ.get("IRIS_PROCESSED_BITRATE", "12M")
PROCESSED_MAXRATE = os.environ.get("IRIS_PROCESSED_MAXRATE", "16M")
PROCESSED_BUFSIZE = os.environ.get("IRIS_PROCESSED_BUFSIZE", "10M")
# Upload processed streams use tighter bitrate by default for smoother remote playback.
UPLOAD_PROCESSED_BITRATE = os.environ.get("IRIS_UPLOAD_PROCESSED_BITRATE", "6M")
UPLOAD_PROCESSED_MAXRATE = os.environ.get("IRIS_UPLOAD_PROCESSED_MAXRATE", "8M")
UPLOAD_PROCESSED_BUFSIZE = os.environ.get("IRIS_UPLOAD_PROCESSED_BUFSIZE", "4M")
PROCESSED_NVENC_PRESET = os.environ.get("IRIS_PROCESSED_NVENC_PRESET", "p5")
PROCESSED_FPS_BOOST = float(os.environ.get("IRIS_PROCESSED_FPS_BOOST", "2"))
RAW_BITRATE = os.environ.get("IRIS_RAW_BITRATE", "3M")
RAW_MAXRATE = os.environ.get("IRIS_RAW_MAXRATE", "4M")
RAW_BUFSIZE = os.environ.get("IRIS_RAW_BUFSIZE", "2M")
PERSIST_ACTIVE_SOURCES = os.environ.get("IRIS_PERSIST_ACTIVE_SOURCES", "0") == "1"
MAX_RTSP_STREAMS = int(os.environ.get("IRIS_MAX_RTSP_STREAMS", "4"))
MAX_UPLOAD_STREAMS = int(os.environ.get("IRIS_MAX_UPLOAD_STREAMS", "2"))
USE_NVENC = os.environ.get("IRIS_USE_NVENC", "1") == "1"
RAW_USE_NVENC = os.environ.get("IRIS_RAW_USE_NVENC", "0") == "1"
# For uploaded files, publish raw via direct ffmpeg->MediaMTX low-latency transcode path.
UPLOAD_RAW_PASSTHROUGH = os.environ.get("IRIS_UPLOAD_RAW_PASSTHROUGH", "0") == "1"
UPLOAD_REALTIME = os.environ.get("IRIS_UPLOAD_REALTIME", "1") == "1"
MEMORY_GUARD_ENABLED = os.environ.get("IRIS_MEMORY_GUARD_ENABLED", "1") == "1"
MEMORY_GUARD_AVAIL_MB = int(os.environ.get("IRIS_MEMORY_GUARD_AVAIL_MB", "1024"))
MEMORY_GUARD_COOLDOWN_SECONDS = int(os.environ.get("IRIS_MEMORY_GUARD_COOLDOWN_SECONDS", "10"))
MEMORY_GUARD_POLL_SECONDS = float(os.environ.get("IRIS_MEMORY_GUARD_POLL_SECONDS", "2"))
VERBOSE_OVERLAY_INIT_LOGS = os.environ.get("IRIS_VERBOSE_OVERLAY_INIT_LOGS", "0") == "1"
PROCESSED_JPEG_ENCODE_PARAMS = [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY]
RAW_JPEG_ENCODE_PARAMS = [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY]

# Active mode tracking
active_mode: Optional[str] = None
inference_pause_lock = threading.Lock()
inference_pause_until: float = 0.0
inference_pause_reason: str = ""

MODE_PROFILES = {
    "congestion": {
        "overlay": {"heatmap": True, "heatmap_full": False, "heatmap_trails": True, "trails": False, "bboxes": False},
        "confidence": 0.20,
    },
    "vehicle": {
        "overlay": {"heatmap": False, "heatmap_full": False, "heatmap_trails": False, "trails": False, "bboxes": True, "bbox_label": "class"},
        "confidence": 0.25,
    },
    "flow": {
        "overlay": {"heatmap": False, "heatmap_full": False, "heatmap_trails": False, "trails": True, "bboxes": True, "bbox_label": "speed"},
        "confidence": 0.20,
    },
    "vision": {
        "overlay": {"heatmap": False, "heatmap_full": False, "heatmap_trails": False, "trails": False, "bboxes": False, "bbox_label": "class"},
        "confidence": 0.20,
    },
    "forensics": {
        "overlay": {"heatmap": False, "heatmap_full": False, "heatmap_trails": False, "trails": False, "bboxes": False, "bbox_label": "class"},
        "confidence": 0.20,
    },
    "crowd": {
        "overlay": {"heatmap": True, "heatmap_full": False, "heatmap_trails": False, "trails": False, "bboxes": False, "bbox_label": "class"},
        "confidence": 0.25,
    },
}
MODE_ALIASES = {
    "forensices": "forensics",
    "iris_forensics": "forensics",
    "iris-vision": "vision",
    "upload": "vehicle",  # uploads use vehicle inference (VCC bboxes); Gemini runs separately
    "realtime-forensics": "forensics",
}


def normalize_mode(raw_mode: Optional[str]) -> Optional[str]:
    if raw_mode is None:
        return None
    mode = str(raw_mode).strip().lower()
    if not mode:
        return None
    mode = MODE_ALIASES.get(mode, mode)
    return mode if mode in MODE_PROFILES else None


def route_mode_overlay(mode: Optional[str]) -> dict:
    normalized = normalize_mode(mode)
    if normalized and normalized in MODE_PROFILES:
        return dict(MODE_PROFILES[normalized]["overlay"])
    return {"heatmap": False, "heatmap_full": False, "heatmap_trails": False, "trails": False, "bboxes": False}


def route_mode_confidence(mode: Optional[str]) -> float:
    normalized = normalize_mode(mode)
    if normalized and normalized in MODE_PROFILES:
        return float(MODE_PROFILES[normalized]["confidence"])
    return 0.15


MODE_CONFIDENCE = {mode: float(cfg.get("confidence", 0.15)) for mode, cfg in MODE_PROFILES.items()}

confidence_lock = threading.Lock()
confidence_settings: Dict[str, float] = {}

start_source_callback = None
start_upload_callback = None

frontend_log_lock = threading.Lock()
frontend_log_file = None

# ── Auth/session state ──

auth_lock = threading.Lock()
auth_sessions: Dict[str, dict] = {}
auth_user_tokens: Dict[str, set] = {}  # username → set of active tokens
AUTH_IDLE_TTL_SECONDS = int(os.environ.get("IRIS_AUTH_IDLE_TTL_SECONDS", "1800"))
AUTH_PUBLIC_PATHS = {
    "/api/login",
    "/api/health",
    "/api/logs/frontend",
    "/api/realtime-forensics/report/pdf",
    "/api/crowd-worker/config",
    "/api/crowd/analysis",
    "/api/crowd/analysis/latest",
    "/api/crowd-live/report/pdf",
    "/api/crowd-live/master-report/pdf",
    "/api/crowd-live/csv",
}

# ── Crowd-worker analysis storage ──
_crowd_worker_analysis: Dict[str, dict] = {}
_crowd_worker_lock = threading.Lock()
_HEATMAP_DIR = os.path.join(os.path.expanduser("~"), "heatmaps")
os.makedirs(_HEATMAP_DIR, exist_ok=True)

# Crowd-worker history for reports (accumulates per-camera data over time)
_CW_DATA_DIR = Path(__file__).resolve().parent / "data" / "crowd_worker"
_CW_DATA_DIR.mkdir(parents=True, exist_ok=True)
_crowd_worker_history: Dict[str, list] = {}  # deviceId → list of analysis snapshots
_crowd_worker_session_start: str = ""
_CW_SEGMENT_INTERVAL = 600  # 10 minutes per segment
_cw_last_segment_time = 0.0
_cw_segments: list = []  # accumulated segment records for reports
_cw_csv_path = _CW_DATA_DIR / f"crowd_events_{datetime.now().strftime('%Y-%m-%d')}.csv"


def _cw_append_csv(data: dict):
    """Append a crowd-worker analysis record to today's CSV."""
    import csv
    global _cw_csv_path
    today = datetime.now().strftime('%Y-%m-%d')
    expected = _CW_DATA_DIR / f"crowd_events_{today}.csv"
    if expected != _cw_csv_path:
        _cw_csv_path = expected
    write_header = not _cw_csv_path.exists()
    fields = ['timestamp', 'deviceId', 'peopleCount', 'yoloCount', 'densityCount',
              'densityLevel', 'congestionLevel', 'overall_risk', 'sentiment',
              'crowd_movement', 'behavior', 'weapon_detected', 'fight_collision_injury',
              'wrongful_activity', 'visibility_score', 'safety_precaution', 'heatmapImageUrl']
    try:
        with open(_cw_csv_path, 'a', newline='') as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
            if write_header:
                w.writeheader()
            w.writerow(data)
    except Exception:
        pass


def _cw_maybe_close_segment():
    """Check if 10 minutes have passed and close the current segment."""
    global _cw_last_segment_time
    now = time.time()
    if _cw_last_segment_time == 0:
        _cw_last_segment_time = now
        return
    if now - _cw_last_segment_time < _CW_SEGMENT_INTERVAL:
        return

    # Close segment
    seg_start_dt = datetime.fromtimestamp(_cw_last_segment_time)
    seg_end_dt = datetime.fromtimestamp(now)
    seg_record = {
        'segment_index': len(_cw_segments) + 1,
        'segment_start': seg_start_dt.strftime('%H:%M:%S'),
        'segment_end': seg_end_dt.strftime('%H:%M:%S'),
        'per_capture_counts': [],
        'per_camera_data': {},
        'analysis': {},
    }
    for device_id, snapshots in _crowd_worker_history.items():
        if not snapshots:
            continue
        counts = [s.get('peopleCount', 0) for s in snapshots]
        cam_name = snapshots[-1].get('deviceId', device_id).replace('camera_', '').replace('_', '.')
        # Find the heatmap image path from the latest snapshot
        heatmap_url = snapshots[-1].get('heatmapImageUrl', '')
        heatmap_path = os.path.join(_HEATMAP_DIR, heatmap_url.split('/heatmaps/')[-1]) if '/heatmaps/' in heatmap_url else ''

        seg_record['per_capture_counts'].extend(counts)
        seg_record['per_camera_data'][device_id] = {
            'camera_name': cam_name,
            'counts': counts,
            'avg_count': round(sum(counts) / len(counts), 1) if counts else 0,
            'peak_count': max(counts) if counts else 0,
            'heatmap_path': heatmap_path if os.path.exists(heatmap_path) else '',
            'heatmap_frame_count': snapshots[-1].get('peopleCount', 0),
            'num_cameras': 1,
        }
        # Use latest Gemini analysis
        latest = snapshots[-1]
        seg_record['analysis'] = {
            'crowd_movement': latest.get('crowd_movement', ''),
            'crowd_density': latest.get('densityLevel', 'LOW'),
            'sentiment': latest.get('sentiment', 'NEUTRAL'),
            'overall_risk': latest.get('overall_risk', 'LOW'),
            'safety_precaution': latest.get('safety_precaution', ''),
            'weapon_detected': latest.get('weapon_detected', 'NO'),
            'fight_collision_injury': latest.get('fight_collision_injury', 'NO'),
            'wrongful_activity': latest.get('wrongful_activity', 'NO'),
        }

    _cw_segments.append(seg_record)
    # Clear history for next segment
    _crowd_worker_history.clear()
    _cw_last_segment_time = now
    print(f"[CW-REPORT] Segment {seg_record['segment_index']} closed: {seg_record['segment_start']} - {seg_record['segment_end']}")

DEFAULT_CORS_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:5173",
    "https://iriscmd.stagingbot.xyz",
    "https://drone.magicboxhub.net",
]
_cors_origins_raw = os.environ.get("IRIS_CORS_ORIGINS", ",".join(DEFAULT_CORS_ORIGINS))
CORS_ORIGINS = [o.strip() for o in _cors_origins_raw.split(",") if o.strip()]
MAGICBOXHUB_ORIGIN_RE = re.compile(r"^https://([a-zA-Z0-9-]+\.)*magicboxhub\.net$")


def _client_ip_from_request(request: Request) -> str:
    xfwd = request.headers.get("x-forwarded-for", "").strip()
    if xfwd:
        ip = xfwd.split(",")[0].strip()
        if ip:
            return ip
    xreal = request.headers.get("x-real-ip", "").strip()
    if xreal:
        return xreal
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _extract_bearer_token(request: Request) -> str:
    authz = request.headers.get("authorization", "").strip()
    if authz.lower().startswith("bearer "):
        return authz[7:].strip()
    fallback = request.headers.get("x-auth-token", "").strip()
    if fallback:
        return fallback
    q = request.query_params.get("auth_token", "").strip()
    return q


def _inference_pause_remaining_seconds(now: Optional[float] = None) -> float:
    if now is None:
        now = time.time()
    with inference_pause_lock:
        return max(0.0, inference_pause_until - now)


def _set_inference_pause(seconds: int, reason: str) -> float:
    global inference_pause_until, inference_pause_reason
    until = time.time() + max(1, int(seconds))
    with inference_pause_lock:
        if until > inference_pause_until:
            inference_pause_until = until
            inference_pause_reason = reason
        return inference_pause_until


def _ensure_inference_can_start() -> None:
    remaining = _inference_pause_remaining_seconds()
    if remaining <= 0:
        return
    wait_s = int(math.ceil(remaining))
    with inference_pause_lock:
        reason = inference_pause_reason or "memory recovery"
    raise HTTPException(503, f"Inference temporarily paused ({reason}). Retry in {wait_s}s.")


def _purge_stale_auth(now: Optional[float] = None):
    if now is None:
        now = time.time()
    purge_expired_sessions(AUTH_IDLE_TTL_SECONDS, int(now))
    stale_tokens = []
    for token, sess in list(auth_sessions.items()):
        last_seen = float(sess.get("last_seen", 0.0))
        if now - last_seen > AUTH_IDLE_TTL_SECONDS:
            stale_tokens.append((token, sess.get("username")))
    for token, username in stale_tokens:
        auth_sessions.pop(token, None)
        if username and username in auth_user_tokens:
            auth_user_tokens[username].discard(token)
            if not auth_user_tokens[username]:
                auth_user_tokens.pop(username, None)


def _is_origin_allowed(origin: str) -> bool:
    if not origin:
        return False
    if origin in CORS_ORIGINS:
        return True
    return bool(MAGICBOXHUB_ORIGIN_RE.match(origin))


def _cors_headers_for_request(request: Request) -> Dict[str, str]:
    origin = request.headers.get("origin", "").strip()
    base_headers = {
        "Access-Control-Allow-Methods": "GET,POST,PUT,PATCH,DELETE,OPTIONS",
        "Access-Control-Allow-Headers": "Authorization,Content-Type,X-Client-Tab,X-Auth-Token",
        "Access-Control-Allow-Credentials": "true",
    }
    if _is_origin_allowed(origin):
        return {
            **base_headers,
            "Access-Control-Allow-Origin": origin,
            "Vary": "Origin",
        }
    return {
        **base_headers,
        "Access-Control-Allow-Origin": "*",
    }


# ── RTSP config helpers ──

def _extract_rtsp_links_from_mediamtx_cfg() -> List[str]:
    """Best-effort fallback: derive RTSP source URLs from MediaMTX path config."""
    mediamtx_cfg_path = Path("config/mediamtx.yml")
    if not mediamtx_cfg_path.exists():
        return []
    try:
        with open(mediamtx_cfg_path, "r") as f:
            cfg = yaml.safe_load(f) or {}
    except Exception:
        return []

    paths = cfg.get("paths", {}) if isinstance(cfg, dict) else {}
    if not isinstance(paths, dict):
        return []

    out = []
    for _, p in paths.items():
        if not isinstance(p, dict):
            continue
        src = str(p.get("source", "")).strip()
        if src.startswith("rtsp://"):
            out.append(src)
    return out


def _normalize_rtsp_config(cfg: dict | None) -> dict:
    if not isinstance(cfg, dict):
        cfg = {}
    links = cfg.get("rtsp_links")
    active = cfg.get("active_sources")
    overlays_cfg = cfg.get("overlays")

    if not isinstance(links, list):
        links = []
    if not isinstance(active, list):
        active = []
    if not isinstance(overlays_cfg, dict):
        overlays_cfg = {}

    # Auto-heal: if rtsp_links disappeared, recover from mediamtx path sources.
    if not links:
        recovered = _extract_rtsp_links_from_mediamtx_cfg()
        if recovered:
            print(f"[CONFIG] Recovered {len(recovered)} rtsp_links from config/mediamtx.yml")
            links = recovered

    out = {"rtsp_links": links, "active_sources": active, "overlays": overlays_cfg}
    # Preserve crowd_cameras and groups if present
    if isinstance(cfg.get("crowd_cameras"), list):
        out["crowd_cameras"] = cfg["crowd_cameras"]
    if isinstance(cfg.get("crowd_camera_groups"), list):
        out["crowd_camera_groups"] = cfg["crowd_camera_groups"]
    return out


def load_rtsp_config():
    if not RTSP_CONFIG_PATH.exists():
        return {"rtsp_links": [], "active_sources": [], "overlays": {}}
    with open(RTSP_CONFIG_PATH, "r") as f:
        raw = yaml.safe_load(f) or {}
    return _normalize_rtsp_config(raw)

def save_rtsp_config(cfg):
    cfg = _normalize_rtsp_config(cfg)
    with open(RTSP_CONFIG_PATH, "w") as f:
        yaml.safe_dump(cfg, f, default_flow_style=False, sort_keys=False)

def source_display_name(url: str) -> str:
    try:
        from urllib.parse import urlparse
        host = urlparse(url).hostname or ""
        if host:
            return "cam_" + host.replace(".", "_")
    except Exception:
        pass
    return url.rstrip("/").split("/")[-1]

def mask_rtsp(url: str) -> str:
    if "@" in url and "://" in url:
        scheme, rest = url.split("://", 1)
        creds, host = rest.split("@", 1)
        user = creds.split(":")[0]
        return f"{scheme}://{user}:****@{host}"
    return url


def _write_frontend_log(entry: FrontendLogEntry):
    global frontend_log_file
    with frontend_log_lock:
        if frontend_log_file is None:
            log_path = env_log_path("frontend") or new_log_path("frontend")
            frontend_log_file = open(log_path, "a", buffering=1)
        ts = entry.ts or time.strftime("%Y-%m-%dT%H:%M:%S")
        level = (entry.level or "info").upper()
        msg = entry.message or ""
        context = ""
        if entry.context:
            try:
                context = " " + json.dumps(entry.context, separators=(",", ":"), ensure_ascii=True)
            except Exception:
                context = ""
        frontend_log_file.write(f"{ts} {level} {msg}{context}\n")


# ── Mediamtx helpers ──

def _mediamtx_add_path_sync(name: str, rtsp_url: str):
    try:
        http_requests.delete(f"{MEDIAMTX_API}/v3/config/paths/delete/{name}", timeout=2)
    except Exception:
        pass
    try:
        r = http_requests.post(
            f"{MEDIAMTX_API}/v3/config/paths/add/{name}",
            json={"source": rtsp_url, "sourceOnDemand": False},
            timeout=3,
        )
        print(f"[MEDIAMTX] Added path {name} (status={r.status_code})")
    except Exception as e:
        print(f"[MEDIAMTX] Failed to add path {name}: {e}")

def _mediamtx_remove_path_sync(name: str):
    try:
        r = http_requests.delete(f"{MEDIAMTX_API}/v3/config/paths/delete/{name}", timeout=2)
        print(f"[MEDIAMTX] Deleted path {name} (status={r.status_code})")
    except Exception as e:
        print(f"[MEDIAMTX] Failed to delete path {name}: {e}")

def mediamtx_add_path(name: str, rtsp_url: str):
    threading.Thread(target=_mediamtx_add_path_sync, args=(name, rtsp_url), daemon=True).start()

def mediamtx_remove_path(name: str):
    threading.Thread(target=_mediamtx_remove_path_sync, args=(name,), daemon=True).start()


# ── FFmpeg publisher management ──

def _wait_for_frame(name: str, timeout: float = 5.0):
    deadline = time.time() + timeout
    with frame_lock:
        last_seq = frame_sequences.get(name, 0)
    while time.time() < deadline:
        with frame_lock:
            if frame_sequences.get(name, 0) > last_seq:
                data = frame_buffer.get(name)
                if data:
                    return data
            frame_lock.wait(timeout=0.3)
    return None

def _wait_for_frame_bgr(name: str, timeout: float = 5.0):
    deadline = time.time() + timeout
    with frame_lock:
        last_seq = frame_sequences.get(name, 0)
    while time.time() < deadline:
        with frame_lock:
            if frame_sequences.get(name, 0) > last_seq:
                frame = frame_bgr_buffer.get(name)
                if frame is not None:
                    return frame
            frame_lock.wait(timeout=0.3)
    return None

def _wait_for_raw_bgr_frame(name: str, timeout: float = 5.0):
    deadline = time.time() + timeout
    with raw_frame_lock:
        last_seq = raw_frame_sequences.get(name, 0)
    while time.time() < deadline:
        with raw_frame_lock:
            if raw_frame_sequences.get(name, 0) > last_seq:
                frame = raw_frame_bgr_buffer.get(name)
                if frame is not None:
                    return frame
            raw_frame_lock.wait(timeout=0.3)
    return None


def _target_output_fps(name: str, processed: bool = True) -> float:
    fps = float(PROCESSED_FPS)
    is_upload = False
    with upload_sources_lock:
        info = upload_sources.get(name)
        if info:
            is_upload = True
            src_fps = float(info.get("source_fps") or 0.0)
            if 1.0 <= src_fps <= 120.0:
                fps = src_fps
    # Keep uploads at source FPS to avoid duplicated frames/jitter in processed playback.
    if processed and not is_upload:
        fps += max(0.0, PROCESSED_FPS_BOOST)
    return min(60.0, max(5.0, fps))


def _ffmpeg_writer_worker(
    name: str,
    proc: subprocess.Popen,
    stop_event: threading.Event,
    width: int,
    height: int,
    out_fps: float,
):
    last_seq = -1
    last_frame = None
    frame_interval = 1.0 / out_fps
    next_send_time = time.monotonic()
    print(f"[FFMPEG] Writer thread started for {name} (w={width} h={height} fps={out_fps})")
    while not stop_event.is_set():
        # Rate-limit to match PROCESSED_FPS so FFmpeg timestamps stay in sync.
        now = time.monotonic()
        sleep_dur = next_send_time - now
        if sleep_dur > 0:
            time.sleep(sleep_dur)
        next_send_time = max(time.monotonic(), next_send_time) + frame_interval

        with frame_lock:
            seq = frame_sequences.get(name, 0)
            frame = frame_bgr_buffer.get(name)

        if stop_event.is_set():
            break
        if frame is not None and seq > last_seq:
            if frame.shape[1] != width or frame.shape[0] != height:
                print(f"[FFMPEG] Frame size changed for {name} ({width}x{height} -> {frame.shape[1]}x{frame.shape[0]}), restarting.")
                break
            last_frame = frame
            last_seq = seq

        if last_frame is None:
            continue

        try:
            if proc.stdin is None:
                print(f"[FFMPEG] Writer stdin unavailable for {name}, stopping writer loop.")
                break
            proc.stdin.write(last_frame.tobytes())
            count = ffmpeg_frame_sequences.get(name, 0) + 1
            ffmpeg_frame_sequences[name] = count
            if count == 1:
                print(f"[FFMPEG] First frame written for {name} (seq={last_seq})")
        except BrokenPipeError:
            print(f"[FFMPEG] BrokenPipe for {name}, stopping writer loop.")
            break
        except Exception as e:
            print(f"[FFMPEG] Writer error for {name}: {e}")
            break

    try:
        if proc.stdin:
            proc.stdin.close()
    except Exception:
        pass

def _raw_ffmpeg_writer_worker(
    name: str,
    proc: subprocess.Popen,
    stop_event: threading.Event,
    width: int,
    height: int,
    out_fps: float,
):
    last_seq = -1
    frame_interval = 1.0 / out_fps
    next_send_time = time.monotonic()
    while not stop_event.is_set():
        with raw_frame_lock:
            while raw_frame_sequences.get(name, 0) <= last_seq and not stop_event.is_set():
                raw_frame_lock.wait(timeout=0.3)
            seq = raw_frame_sequences.get(name, 0)
            frame = raw_frame_bgr_buffer.get(name)

        if stop_event.is_set():
            break
        if frame is None or seq <= last_seq:
            time.sleep(0.01)
            continue

        now = time.monotonic()
        sleep_dur = next_send_time - now
        if sleep_dur > 0:
            time.sleep(sleep_dur)
        next_send_time = max(time.monotonic(), next_send_time) + frame_interval

        last_seq = seq
        if frame.shape[1] != width or frame.shape[0] != height:
            print(f"[FFMPEG RAW] Frame size changed for {name} ({width}x{height} -> {frame.shape[1]}x{frame.shape[0]}), restarting.")
            break

        try:
            proc.stdin.write(frame.tobytes())
            proc.stdin.flush()
        except BrokenPipeError:
            break
        except Exception as e:
            print(f"[FFMPEG RAW] Writer error for {name}: {e}")
            break

    try:
        if proc.stdin:
            proc.stdin.close()
    except Exception:
        pass


def _start_ffmpeg_publisher(name: str, wait_timeout: float = 8.0) -> bool:
    with ffmpeg_state_lock:
        if name in ffmpeg_starting:
            return False
        existing = ffmpeg_publishers.get(name)
        if existing is not None and existing.poll() is None:
            return True
        ffmpeg_starting.add(name)

    # Keep the in-flight "starting" marker while replacing any stale publisher
    # so concurrent start attempts can't race and spawn duplicates.
    _stop_ffmpeg_publisher(name, clear_starting=False)
    rtsp_url = f"{MEDIAMTX_PUBLISH_BASE}/processed_{name}"

    frame = _wait_for_frame_bgr(name, timeout=wait_timeout)
    if frame is None:
        print(f"[FFMPEG] No frames available for {name}, skipping publisher start.")
        with ffmpeg_state_lock:
            ffmpeg_starting.discard(name)
        return False

    height, width = frame.shape[:2]
    out_fps = _target_output_fps(name, processed=True)
    gop = max(1, int(round(out_fps * KEYFRAME_INTERVAL_SEC)))
    with upload_sources_lock:
        is_upload_stream = name in upload_sources
    target_bitrate = UPLOAD_PROCESSED_BITRATE if is_upload_stream else PROCESSED_BITRATE
    target_maxrate = UPLOAD_PROCESSED_MAXRATE if is_upload_stream else PROCESSED_MAXRATE
    target_bufsize = UPLOAD_PROCESSED_BUFSIZE if is_upload_stream else PROCESSED_BUFSIZE
    if USE_NVENC:
        encoder_args = [
            "-c:v", "h264_nvenc",
            "-preset", PROCESSED_NVENC_PRESET,
            "-tune", "ll",
            "-rc", "cbr",
            "-b:v", target_bitrate,
            "-maxrate", target_maxrate,
            "-bufsize", target_bufsize,
            "-spatial-aq", "1",
            "-aq-strength", "8",
            "-forced-idr", "1",
        ]
    else:
        encoder_args = [
            "-c:v", "libx264",
            "-preset", "ultrafast",
            "-tune", "zerolatency",
            "-b:v", target_bitrate,
            "-maxrate", target_maxrate,
            "-bufsize", target_bufsize,
            "-x264-params", f"keyint={gop}:min-keyint={gop}:scenecut=0",
        ]

    cmd = [
        "ffmpeg",
        "-nostdin",
        "-loglevel", "warning",
        "-fflags", "nobuffer",
        "-flags", "low_delay",
        "-f", "rawvideo",
        "-pix_fmt", "bgr24",
        "-s", f"{width}x{height}",
        "-r", str(out_fps),
        "-i", "pipe:0",
        "-vf", "format=yuv420p",
        *encoder_args,
        "-g", str(gop),
        "-bf", "0",
        "-threads", "1",
        "-f", "rtsp",
        "-rtsp_transport", "tcp",
        rtsp_url,
    ]

    try:
        ffmpeg_log_path = env_log_path("ffmpeg") or (new_log_path("ffmpeg").parent / "ffmpeg.log")
        ffmpeg_log_path.parent.mkdir(parents=True, exist_ok=True)
        log_file = open(ffmpeg_log_path, "a", buffering=1)
        ffmpeg_log_files[name] = log_file
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=log_file,
            stderr=log_file,
            start_new_session=True,
        )
        ffmpeg_publishers[name] = proc
        stop_event = threading.Event()
        ffmpeg_stop_events[name] = stop_event
        thread = threading.Thread(
            target=_ffmpeg_writer_worker,
            args=(name, proc, stop_event, width, height, out_fps),
            daemon=True,
        )
        ffmpeg_writer_threads[name] = thread
        thread.start()
        print(f"[FFMPEG] Started publisher for {name} -> processed_{name} (pid={proc.pid})")
        with ffmpeg_state_lock:
            ffmpeg_starting.discard(name)
        return True
    except Exception as e:
        print(f"[FFMPEG] Failed to start publisher for {name}: {e}")
        with ffmpeg_state_lock:
            ffmpeg_starting.discard(name)
        return False

def _start_raw_ffmpeg_publisher(name: str, wait_timeout: float = 10.0) -> bool:
    _stop_raw_ffmpeg_publisher(name)
    rtsp_url = f"{MEDIAMTX_PUBLISH_BASE}/{name}"

    frame = _wait_for_raw_bgr_frame(name, timeout=wait_timeout)
    if frame is None:
        print(f"[FFMPEG RAW] No frames available for {name}, skipping raw publisher start.")
        return False

    height, width = frame.shape[:2]
    out_fps = _target_output_fps(name, processed=False)
    gop = max(1, int(round(out_fps * KEYFRAME_INTERVAL_SEC)))
    if RAW_USE_NVENC:
        encoder_args = [
            "-c:v", "h264_nvenc",
            "-preset", "p1",
            "-tune", "ll",
            "-rc", "cbr",
            "-b:v", RAW_BITRATE,
            "-maxrate", RAW_MAXRATE,
            "-bufsize", RAW_BUFSIZE,
            "-forced-idr", "1",
        ]
    else:
        encoder_args = [
            "-c:v", "libx264",
            "-preset", "ultrafast",
            "-tune", "zerolatency",
            "-b:v", RAW_BITRATE,
            "-maxrate", RAW_MAXRATE,
            "-bufsize", RAW_BUFSIZE,
            "-x264-params", f"keyint={gop}:min-keyint={gop}:scenecut=0",
        ]

    cmd = [
        "ffmpeg",
        "-nostdin",
        "-loglevel", "warning",
        "-fflags", "nobuffer",
        "-flags", "low_delay",
        "-f", "rawvideo",
        "-pix_fmt", "bgr24",
        "-s", f"{width}x{height}",
        "-r", str(out_fps),
        "-i", "pipe:0",
        "-vf", "format=yuv420p",
        *encoder_args,
        "-g", str(gop),
        "-bf", "0",
        "-threads", "1",
        "-f", "rtsp",
        "-rtsp_transport", "tcp",
        rtsp_url,
    ]

    try:
        ffmpeg_log_path = env_log_path("ffmpeg") or (new_log_path("ffmpeg").parent / "ffmpeg.log")
        ffmpeg_log_path.parent.mkdir(parents=True, exist_ok=True)
        log_file = open(ffmpeg_log_path, "a", buffering=1)
        raw_ffmpeg_log_files[name] = log_file
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=log_file,
            stderr=log_file,
            start_new_session=True,
        )
        raw_ffmpeg_publishers[name] = proc
        stop_event = threading.Event()
        raw_ffmpeg_stop_events[name] = stop_event
        thread = threading.Thread(
            target=_raw_ffmpeg_writer_worker,
            args=(name, proc, stop_event, width, height, out_fps),
            daemon=True,
        )
        raw_ffmpeg_writer_threads[name] = thread
        thread.start()
        print(f"[FFMPEG RAW] Started publisher for {name} -> {name} (pid={proc.pid})")
        return True
    except Exception as e:
        print(f"[FFMPEG RAW] Failed to start publisher for {name}: {e}")
        return False

def _start_raw_ffmpeg_publisher_async(name: str):
    def _run():
        for _ in range(3):
            if _start_raw_ffmpeg_publisher(name):
                return
            time.sleep(1.5)
    threading.Thread(target=_run, daemon=True).start()

def _start_upload_raw_passthrough_publisher(name: str, file_path: str) -> bool:
    """Publish upload file directly to MediaMTX raw path with low-latency transcode."""
    _stop_raw_ffmpeg_publisher(name)
    rtsp_url = f"{MEDIAMTX_PUBLISH_BASE}/{name}"
    gop = max(1, PROCESSED_FPS * KEYFRAME_INTERVAL_SEC)

    cmd = [
        "ffmpeg",
        "-nostdin",
        "-loglevel", "warning",
        "-re",
        "-stream_loop", "-1",
        "-fflags", "nobuffer",
        "-flags", "low_delay",
        "-i", file_path,
        "-map", "0:v:0",
        "-an",
        "-c:v", "libx264",
        "-preset", "ultrafast",
        "-tune", "zerolatency",
        "-pix_fmt", "yuv420p",
        "-g", str(gop),
        "-keyint_min", str(gop),
        "-sc_threshold", "0",
        "-b:v", RAW_BITRATE,
        "-maxrate", RAW_MAXRATE,
        "-bufsize", RAW_BUFSIZE,
        "-f", "rtsp",
        "-rtsp_transport", "tcp",
        rtsp_url,
    ]

    try:
        ffmpeg_log_path = env_log_path("ffmpeg") or (new_log_path("ffmpeg").parent / "ffmpeg.log")
        ffmpeg_log_path.parent.mkdir(parents=True, exist_ok=True)
        log_file = open(ffmpeg_log_path, "a", buffering=1)
        raw_ffmpeg_log_files[name] = log_file
        proc = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=log_file,
            stderr=log_file,
            start_new_session=True,
        )
        raw_ffmpeg_publishers[name] = proc
        stop_event = threading.Event()
        raw_ffmpeg_stop_events[name] = stop_event
        print(f"[FFMPEG RAW] Started upload low-latency transcode for {name} -> {name} (pid={proc.pid})")
        return True
    except Exception as e:
        print(f"[FFMPEG RAW] Failed upload low-latency transcode for {name}: {e}")
        return False

def _start_upload_raw_passthrough_async(name: str, file_path: str):
    def _run():
        for _ in range(3):
            if _start_upload_raw_passthrough_publisher(name, file_path):
                return
            time.sleep(1.0)
    threading.Thread(target=_run, daemon=True).start()

def _start_ffmpeg_publisher_async(name: str):
    def _run():
        for _ in range(5):
            if _start_ffmpeg_publisher(name):
                return
            time.sleep(1.0)
    threading.Thread(target=_run, daemon=True).start()

def _stop_ffmpeg_publisher(name: str, clear_starting: bool = True):
    if clear_starting:
        with ffmpeg_state_lock:
            ffmpeg_starting.discard(name)
    ffmpeg_next_start.pop(name, None)
    ffmpeg_frame_sequences.pop(name, None)
    stop_event = ffmpeg_stop_events.pop(name, None)
    if stop_event:
        stop_event.set()

    proc = ffmpeg_publishers.pop(name, None)
    if proc:
        try:
            proc.terminate()
            proc.wait(timeout=3)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

    thread = ffmpeg_writer_threads.pop(name, None)
    if thread:
        try:
            thread.join(timeout=1)
        except Exception:
            pass

    log_file = ffmpeg_log_files.pop(name, None)
    if log_file:
        try:
            log_file.close()
        except Exception:
            pass
        print(f"[FFMPEG] Stopped publisher for {name}")

def _stop_raw_ffmpeg_publisher(name: str):
    stop_event = raw_ffmpeg_stop_events.pop(name, None)
    if stop_event:
        stop_event.set()

    proc = raw_ffmpeg_publishers.pop(name, None)
    if proc:
        try:
            proc.terminate()
            proc.wait(timeout=3)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

    thread = raw_ffmpeg_writer_threads.pop(name, None)
    if thread:
        try:
            thread.join(timeout=1)
        except Exception:
            pass

    log_file = raw_ffmpeg_log_files.pop(name, None)
    if log_file:
        try:
            log_file.close()
        except Exception:
            pass
        print(f"[FFMPEG RAW] Stopped publisher for {name}")

_ffmpeg_monitor_sources: set = set()
_ffmpeg_monitor_thread: threading.Thread = None
_ffmpeg_monitor_stop = threading.Event()

def _is_source_name_active(name: str) -> bool:
    cfg = load_rtsp_config()
    links = cfg.get("rtsp_links", [])
    with source_lock:
        for idx in running_sources.keys():
            if 1 <= idx <= len(links) and source_display_name(links[idx - 1]) == name:
                return True
    with upload_sources_lock:
        if name in upload_sources:
            return True
    return False

def _ffmpeg_monitor_worker():
    while not _ffmpeg_monitor_stop.is_set():
        try:
            for name in list(_ffmpeg_monitor_sources):
                with ffmpeg_state_lock:
                    proc = ffmpeg_publishers.get(name)
                now = time.time()
                next_at = ffmpeg_next_start.get(name, 0)

                if proc is None:
                    if not _is_source_name_active(name):
                        _ffmpeg_monitor_sources.discard(name)
                        ffmpeg_next_start.pop(name, None)
                        continue
                    if now >= next_at:
                        ok = _start_ffmpeg_publisher(name)
                        if not ok:
                            # Keep retries aggressive during startup so uploads surface processed
                            # overlays quickly instead of waiting multiple seconds per attempt.
                            ffmpeg_next_start[name] = now + 0.5
                        else:
                            ffmpeg_next_start.pop(name, None)
                    continue

                poll = proc.poll()
                if poll is not None:
                    source_active = _is_source_name_active(name)

                    if source_active:
                        print(f"[FFMPEG] Publisher for {name} exited (code={poll}), restarting...")
                        _stop_ffmpeg_publisher(name)
                        ffmpeg_next_start[name] = now + 0.4
                    else:
                        _ffmpeg_monitor_sources.discard(name)
        except Exception as e:
            print(f"[FFMPEG Monitor] Error: {e}")

        _ffmpeg_monitor_stop.wait(0.35)

def _ensure_ffmpeg_monitor():
    global _ffmpeg_monitor_thread
    if _ffmpeg_monitor_thread is None or not _ffmpeg_monitor_thread.is_alive():
        _ffmpeg_monitor_stop.clear()
        _ffmpeg_monitor_thread = threading.Thread(target=_ffmpeg_monitor_worker, daemon=True)
        _ffmpeg_monitor_thread.start()
        print("[FFMPEG] Monitor thread started")

def _stop_all_ffmpeg_publishers():
    for name in list(ffmpeg_publishers.keys()):
        _stop_ffmpeg_publisher(name)
    for name in list(raw_ffmpeg_publishers.keys()):
        _stop_raw_ffmpeg_publisher(name)


# ── Overlay / metrics / frame management ──

def ensure_overlay(name: str):
    with overlay_lock:
        if name in overlays:
            existing = overlays[name]
            # Ensure all required keys exist
            if isinstance(existing, dict) and "trails" in existing and "heatmap" in existing and "bboxes" in existing:
                if "heatmap_full" not in existing:
                    existing["heatmap_full"] = existing.get("heatmap", True)
                if "heatmap_trails" not in existing:
                    existing["heatmap_trails"] = existing.get("heatmap", True)
                if "confidence" not in existing:
                    existing["confidence"] = 0.15
                return
        
        # Initialize with defaults
        default_state = {
            "heatmap": True,
            "heatmap_full": True,
            "heatmap_trails": True,
            "trails": True,
            "bboxes": True,
            "confidence": 0.15,
        }
        overlays[name] = default_state
        if VERBOSE_OVERLAY_INIT_LOGS:
            print(f"[OVERLAY] Initialized {name}: {default_state}")

def get_overlay_state(stream: str) -> Dict[str, bool]:
    ensure_overlay(stream)
    with overlay_lock:
        return dict(overlays.get(stream, {
            "heatmap": True, 
            "heatmap_full": True, 
            "heatmap_trails": True, 
            "trails": True, 
            "bboxes": True
        }))

def _resolve_source_mode(name: str, source_metrics: Optional[dict] = None) -> Optional[str]:
    """Resolve mode with per-source precedence before falling back to global mode."""
    with upload_sources_lock:
        upload_info = upload_sources.get(name)
        if upload_info and upload_info.get("mode"):
            return normalize_mode(upload_info.get("mode")) or upload_info.get("mode")
    with overlay_lock:
        overlay_mode = (overlays.get(name) or {}).get("active_mode")
        if overlay_mode:
            return normalize_mode(overlay_mode) or overlay_mode
    if isinstance(source_metrics, dict):
        metric_mode = source_metrics.get("mode")
        if metric_mode:
            return normalize_mode(metric_mode) or metric_mode
    return normalize_mode(active_mode) or active_mode


def _to_json_safe(value):
    """Convert nested metric values to JSON-safe builtins."""
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return 0.0
        return value
    if isinstance(value, dict):
        return {str(k): _to_json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_to_json_safe(v) for v in value]
    # numpy / torch scalars
    try:
        if hasattr(value, "item"):
            return _to_json_safe(value.item())
    except Exception:
        pass
    return str(value)


def _default_metrics_for_mode(mode: Optional[str]) -> Dict[str, object]:
    base = {
        "fps": 0.0,
        "detection_count": 0,
        "current_detection_count": 0,
    }
    if mode == "vehicle":
        base.update({
            "class_counts": {},
            "vehicle_total_count": 0,
            "vehicle_class_totals": {},
            "state_counts": {"moving": 0, "stopped": 0, "abnormal": 0},
            "behavior_counts": {"stable": 0, "start_stop": 0, "erratic": 0},
            "type_influence": {},
            "attention_list": [],
            "high_impact_vehicles": [],
            "dwell_stats": {"avg_dwell": 0.0, "max_dwell": 0.0, "over_threshold": 0, "longest_vehicles": []},
        })
    elif mode in ("flow", "congestion"):
        base.update({
            "congestion_index": 0,
            "traffic_density": 0,
            "mobility_index": 0,
            "stalled_pct": 0,
            "slow_pct": 0,
            "medium_pct": 0,
            "fast_pct": 0,
            "class_counts": {},
            "hot_regions": {"active_count": 0, "severity_counts": {"HIGH": 0, "MODERATE": 0, "LOW": 0}, "regions": []},
        })
    elif mode == "crowd":
        base.update({
            "crowd_count": 0,
            "crowd_density": 0,
            "avg_density": 0,
            "peak_density": 0,
            "peak_count": 0,
            "risk_score": 0,
            "density_class": "sparse",
            "operational_status": "MONITOR",
            "zones": [],
            "zone_distribution": {"sparse": 0, "gathering": 0, "dense": 0, "critical": 0},
            "hotspots": [],
            "flow_summary": {"total_inflow": 0, "total_outflow": 0, "net_flow": 0},
            "anomalies": [],
        })
    return base

def update_metrics(stream: str, data: dict):
    # Handle critical workflow signals first
    if data.get("__started__"):
        with source_lock:
            for info in running_sources.values():
                if info.get("name") == stream:
                    info["started_processing"] = True
        with upload_sources_lock:
            info = upload_sources.get(stream)
            if info:
                info["started_processing"] = True
                if not info.get("started_at"):
                    info["started_at"] = time.time()

    if data.get("__finished__"):
        # VCC has finished processing the upload video. Write the sentinel so the
        # forensics _process_job knows VCC is done and can flush its last batch.
        try:
            with upload_sources_lock:
                upload_info = upload_sources.get(stream)
            if upload_info:
                f_job = upload_info.get("forensics_job_id")
                if f_job:
                    frames_dir = forensics_events.get_frames_dir(f_job)
                    if frames_dir:
                        (frames_dir / "vcc_complete.txt").write_text("done")
        except Exception:
            pass
        # Ignore finished flag for uploads to avoid premature completion UI.
        # Reports are triggered manually by the user.
        return

    stream_mode = _resolve_source_mode(stream, data)
    if stream_mode in ("forensics", "vision"):
        # Suppress standard metrics store in report/vision-only modes.
        return

    # Ensure UI can consume metrics as soon as any valid payload arrives.
    with source_lock:
        for info in running_sources.values():
            if info.get("name") == stream:
                info["started_processing"] = True
    with upload_sources_lock:
        info = upload_sources.get(stream)
        if info:
            info["started_processing"] = True
            if not info.get("started_at"):
                info["started_at"] = time.time()

    with metrics_lock:
        payload = _to_json_safe(data or {})
        if not isinstance(payload, dict):
            payload = {}
        if stream_mode and "mode" not in payload:
            payload["mode"] = stream_mode
        mode_for_defaults = payload.get("mode") or stream_mode
        defaults = _default_metrics_for_mode(mode_for_defaults)
        defaults.update(payload)
        metrics[stream] = defaults

def update_frame(stream: str, frame: np.ndarray):
    if frame is None or not isinstance(frame, np.ndarray):
        return

    try:
        ret_enc, buffer = cv2.imencode(".jpg", frame, PROCESSED_JPEG_ENCODE_PARAMS)
        if not ret_enc:
            return
        jpeg_data = buffer.tobytes()
    except Exception:
        return

    # Persist processed frames for forensics event mapping (uploads only).
    # Use the VCC frame sequence counter (not wall-clock time) so the index is
    # deterministic and aligns with the sequential frame numbers that _process_job
    # assigns when it reads the raw video.
    try:
        with upload_sources_lock:
            upload_info = upload_sources.get(stream)
        if upload_info:
            f_job = upload_info.get("forensics_job_id")
            if f_job:
                interval = max(1, int(os.getenv("FORENSICS_FRAME_INTERVAL", str(forensics_events.FORENSICS_FRAME_INTERVAL))))
                # Sequence number before this frame is written (starts at 0).
                frame_index = frame_sequences.get(stream, 0)
                if frame_index % interval == 0:
                    extracted_idx = frame_index // interval
                    frames_dir = forensics_events.get_frames_dir(f_job)
                    if frames_dir:
                        out_path = frames_dir / f"processed_{extracted_idx:05d}.jpg"
                        if not out_path.exists():
                            try:
                                with out_path.open("wb") as f:
                                    f.write(jpeg_data)
                            except Exception:
                                pass
    except Exception:
        pass

    with frame_lock:
        frame_buffer[stream] = jpeg_data
        frame_bgr_buffer[stream] = frame
        frame_sequences[stream] = frame_sequences.get(stream, 0) + 1
        frame_lock.notify_all()

def update_raw_frame(stream: str, frame: np.ndarray):
    if frame is None or not isinstance(frame, np.ndarray):
        return

    try:
        ret_enc, buffer = cv2.imencode(".jpg", frame, RAW_JPEG_ENCODE_PARAMS)
        if not ret_enc:
            return
        jpeg_data = buffer.tobytes()
    except Exception:
        return

    # Mark source/upload active on first raw frame for snappier UI state.
    with source_lock:
        for info in running_sources.values():
            if info.get("name") == stream:
                info["started_processing"] = True
    with upload_sources_lock:
        info = upload_sources.get(stream)
        if info:
            info["started_processing"] = True
            if not info.get("started_at"):
                info["started_at"] = time.time()

    with raw_frame_lock:
        raw_frame_buffer[stream] = jpeg_data
        raw_frame_bgr_buffer[stream] = frame
        raw_frame_sequences[stream] = raw_frame_sequences.get(stream, 0) + 1
        raw_frame_lock.notify_all()

def get_all_metrics():
    with metrics_lock:
        raw_m = metrics.copy()
    m = {k: _to_json_safe(v) for k, v in dict(raw_m).items()}

    # Merge RTSP start times
    with source_lock:
        for info in running_sources.values():
            name = info.get("name")
            start_t = info.get("start_time")
            if name:
                if name not in m:
                    m[name] = {}
                if start_t:
                    m[name]["start_time"] = start_t

    # Merge Upload duration
    with upload_sources_lock:
        for name, info in upload_sources.items():
            if name not in m:
                m[name] = {}
            dur = info.get("duration", 0.0)
            if dur > 0:
                m[name]["total_duration"] = dur

    # Ensure each active source has a complete metrics schema for the UI.
    for name, data in list(m.items()):
        if not isinstance(data, dict):
            data = {}
        mode = _resolve_source_mode(name, data)
        defaults = _default_metrics_for_mode(mode)
        defaults.update(data)
        if mode and "mode" not in defaults:
            defaults["mode"] = mode
        m[name] = defaults

    return m


# ── Alert management ──

def add_alert(source: str, congestion: int, metrics_data: dict, screenshot_data: bytes):
    if congestion < ALERT_TRIGGER_THRESHOLD:
        return None

    with alerts_lock:
        now = time.time()

        last_alert = alert_cooldowns.get(source, 0)
        if now - last_alert < ALERT_COOLDOWN_SECONDS:
            return None

        alert_cooldowns[source] = now

        alert_id = f"{source}_{int(now * 1000)}"
        screenshot_path = ALERTS_DIR / f"{alert_id}.jpg"

        # Alert screenshots must come from processed frames (not raw stream frames).
        # Use queue payload only as a fallback if processed cache is not yet available.
        with frame_lock:
            processed_screenshot = frame_buffer.get(source)
        final_screenshot = (
            bytes(processed_screenshot)
            if isinstance(processed_screenshot, (bytes, bytearray)) and processed_screenshot
            else (bytes(screenshot_data) if isinstance(screenshot_data, (bytes, bytearray)) and screenshot_data else None)
        )
        if not final_screenshot:
            return None

        try:
            with open(screenshot_path, "wb") as f:
                f.write(final_screenshot)
        except Exception as e:
            print(f"[ALERT] Failed to save screenshot: {e}")
            return None

        if congestion >= 85:
            severity = "critical"
        elif congestion >= ALERT_TRIGGER_THRESHOLD:
            severity = "high"
        else:
            severity = "medium"

        alert = {
            "id": alert_id,
            "source": source,
            "severity": severity,
            "congestion": congestion,
            "timestamp": now,
            "time_str": time.strftime("%H:%M:%S"),
            "screenshot": f"/api/alerts/{alert_id}/screenshot",
            "metrics": {
                "congestion_index": metrics_data.get("congestion_index", congestion),
                "traffic_density": metrics_data.get("traffic_density", 0),
                "mobility_index": metrics_data.get("mobility_index", 0),
                "detection_count": metrics_data.get("detection_count", 0),
                "stalled_pct": metrics_data.get("stalled_pct", 0),
                "slow_pct": metrics_data.get("slow_pct", 0),
                "medium_pct": metrics_data.get("medium_pct", 0),
                "fast_pct": metrics_data.get("fast_pct", 0),
            }
        }

        alerts.appendleft(alert)
        print(f"[ALERT] New {severity} alert for {source}: congestion={congestion}%")
        return alert

def load_stored_alerts():
    loaded = 0
    for jpg in sorted(ALERTS_DIR.glob("*.jpg"), key=lambda f: f.stat().st_mtime, reverse=True):
        stem = jpg.stem
        parts = stem.rsplit("_", 1)
        if len(parts) != 2:
            continue
        source = parts[0]
        try:
            ts = int(parts[1]) / 1000.0
        except ValueError:
            continue

        alert = {
            "id": stem,
            "source": source,
            "severity": "medium",
            "congestion": 0,
            "timestamp": ts,
            "time_str": time.strftime("%H:%M:%S", time.localtime(ts)),
            "screenshot": f"/api/alerts/{stem}/screenshot",
            "metrics": {},
        }
        with alerts_lock:
            alerts.append(alert)
        loaded += 1
        if loaded >= 50:
            break
    if loaded:
        print(f"[ALERT] Loaded {loaded} stored alerts from disk")

def get_alerts(limit: int = 20):
    with alerts_lock:
        return list(alerts)[:limit]

def clear_alerts():
    with alerts_lock:
        alerts.clear()
        alert_cooldowns.clear()
    for f in ALERTS_DIR.glob("*.jpg"):
        try:
            f.unlink()
        except:
            pass


# ── Source management ──

def _stop_all_sources():
    global active_mode
    stopped = 0
    import torch
    import gc
    import time

    print(f"[STOP_ALL] Initiating aggressive cleanup (Previous mode: {active_mode})")

    # 1. Stop all primary inference processes
    with source_lock:
        for idx, src in list(running_sources.items()):
            src["stop"].set()
            proc = src.get("process")
            if proc:
                try:
                    proc.terminate()
                    t_start = time.time()
                    while proc.is_alive() and time.time() - t_start < 0.5:
                        time.sleep(0.01)
                    if proc.is_alive():
                        proc.kill()
                except:
                    pass
            stopped += 1
        running_sources.clear()

    # 2. Stop all SAM workers and unload model
    for src_name, info in list(sam_threads.items()):
        print(f"[STOP_ALL] Stopping SAM worker for {src_name}")
        info["stop_event"].set()
    sam_threads.clear()

    try:
        from sam import unload_sam_model
        unload_sam_model()
    except:
        pass

    # 3. Clear SAM analytical results
    with sam_results_lock:
        sam_results.clear()

    # 4. Stop all FFmpeg publishers (streaming)
    _ffmpeg_monitor_sources.clear()
    _stop_all_ffmpeg_publishers()

    # 4b. Stop all upload processes and clear state
    _stop_all_uploads(delete_files=False)

    # 5. Clear ALL shared state/buffers
    with metrics_lock:
        metrics.clear()
    with overlay_lock:
        overlays.clear()
    with frame_lock:
        frame_buffer.clear()
        frame_bgr_buffer.clear()
        frame_sequences.clear()
    with raw_frame_lock:
        raw_frame_buffer.clear()
        raw_frame_bgr_buffer.clear()
        raw_frame_sequences.clear()

    # 6. Force Resource Reclamation
    gc.collect()
    try:
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            print("[STOP_ALL] GPU memory cleared")
    except:
        pass

    # 7. Reset active mode and config
    active_mode = None 
    cfg = load_rtsp_config()
    cfg["active_sources"] = []
    save_rtsp_config(cfg)

    print(f"[STOP_ALL] Cleanup complete. Stopped {stopped} sources.")
    return stopped


# ── FastAPI app ──

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_origin_regex=r"^https://([a-zA-Z0-9-]+\.)*magicboxhub\.net$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

@app.on_event("startup")
def _startup_load_forensics_jobs():
    try:
        forensics_events.load_jobs_from_disk()
    except Exception as e:
        print(f"[FORENSICS] Startup load failed: {e}")

@app.on_event("startup")
def _startup_rt_forensics():
    """Auto-start realtime forensics on backend startup."""
    try:
        result = rt_forensics.start_session()
        print(f"[RT-FORENSICS] Auto-start: {result}")
    except Exception as e:
        print(f"[RT-FORENSICS] Auto-start failed: {e}")

@app.on_event("startup")
def _startup_crowd_auto_start():
    """Auto-start crowd monitoring if cameras are configured."""
    try:
        rt_crowd.auto_start_if_cameras()
    except Exception as e:
        print(f"[RT-CROWD] Auto-start failed: {e}")

@app.on_event("shutdown")
def _shutdown_cancel_forensics_jobs():
    try:
        cancelled = forensics_events.cancel_active_jobs(reason="Server shutting down")
        if cancelled:
            print(f"[FORENSICS] Shutdown: cancelled {len(cancelled)} active job(s): {cancelled}")
    except Exception as e:
        print(f"[FORENSICS] Shutdown cancel failed: {e}")

@app.middleware("http")
async def auth_guard(request: Request, call_next):
    path = request.url.path
    method = request.method.upper()

    # Skip non-API paths and CORS preflight.
    if not path.startswith("/api/") or method == "OPTIONS":
        return await call_next(request)

    # Public API paths.
    if path in AUTH_PUBLIC_PATHS or path.startswith("/api/camera-hls/") or path.startswith("/api/realtime-forensics/event/") or path.startswith("/api/crowd-live/") or path.startswith("/api/heatmaps/") or path.startswith("/api/magicbox-crowd/"):
        return await call_next(request)

    cors_headers = _cors_headers_for_request(request)

    token = _extract_bearer_token(request)
    if not token:
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"}, headers=cors_headers)

    client_ip = _client_ip_from_request(request)
    client_tab = request.headers.get("x-client-tab", "").strip() or request.query_params.get("x_client_tab", "").strip()
    if not client_tab:
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"}, headers=cors_headers)

    with auth_lock:
        _purge_stale_auth()
        sess = auth_sessions.get(token)
        if not sess:
            db_sess = get_session(token)
            if db_sess:
                sess = db_sess
                auth_sessions[token] = dict(sess)
                auth_user_tokens.setdefault(sess.get("username"), set()).add(token)
        if not sess:
            return JSONResponse(status_code=401, content={"detail": "Unauthorized"}, headers=cors_headers)
        if sess.get("tab_id") and sess.get("tab_id") != client_tab:
            return JSONResponse(status_code=401, content={"detail": "Session restricted to one browser tab"}, headers=cors_headers)
        if sess.get("ip") and sess.get("ip") != client_ip:
            return JSONResponse(status_code=401, content={"detail": "Session IP mismatch"}, headers=cors_headers)
        sess["last_seen"] = time.time()
        auth_sessions[token] = sess
        touch_session(token, int(sess["last_seen"]))

    request.state.auth_username = sess.get("username")
    request.state.auth_ip = client_ip
    return await call_next(request)

@app.exception_handler(413)
async def _payload_too_large_handler(request, exc):
    """Return 413 with CORS headers for large file uploads."""
    return JSONResponse(
        status_code=413,
        content={"detail": "File too large. Maximum upload size is 500MB."},
        headers=_cors_headers_for_request(request),
    )

@app.exception_handler(Exception)
async def _global_exception_handler(request, exc):
    """Return 500 with CORS headers so the browser doesn't mask the real error."""
    import traceback
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc)},
        headers=_cors_headers_for_request(request),
    )


# ── Auth endpoints ──

@app.post("/api/login")
def api_login(req: LoginRequest, request: Request):
    result = login_user(req.username, req.password)
    if not result.get("success"):
        raise HTTPException(status_code=401, detail=result.get("error", "Invalid credentials"))

    client_ip = _client_ip_from_request(request)
    client_tab = request.headers.get("x-client-tab", "").strip()
    if not client_tab:
        raise HTTPException(status_code=401, detail="Missing client tab identifier")

    token = uuid.uuid4().hex
    now = time.time()
    with auth_lock:
        auth_user_tokens.setdefault(req.username, set()).add(token)
        auth_sessions[token] = {
            "username": req.username,
            "ip": client_ip,
            "tab_id": client_tab,
            "created_at": now,
            "last_seen": now,
        }
        create_session(token, req.username, client_ip, client_tab, int(now))

    return {"success": True, "username": req.username, "token": token}

@app.post("/api/logout")
def api_logout(request: Request):
    token = _extract_bearer_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")
    with auth_lock:
        sess = auth_sessions.pop(token, None)
        if not sess:
            db_sess = get_session(token)
            if not db_sess:
                raise HTTPException(status_code=401, detail="Unauthorized")
            sess = db_sess
        uname = sess.get("username")
        if uname and uname in auth_user_tokens:
            auth_user_tokens[uname].discard(token)
            if not auth_user_tokens[uname]:
                auth_user_tokens.pop(uname, None)
        delete_session(token)
    return {"success": True}

@app.get("/api/users")
def api_list_users():
    return list_users()

@app.post("/api/users/add")
def api_add_user(req: UserCreateRequest):
    return add_user(req.username, req.password)

@app.post("/api/users/delete")
def api_delete_user(req: UserDeleteRequest):
    return delete_user(req.username)

@app.post("/api/users/change_password")
def api_change_password(req: PasswordChangeRequest):
    return change_password(req.username, req.new_password)


# ── Frontend log endpoint ──

@app.post("/api/logs/frontend")
def api_frontend_log(entry: FrontendLogEntry):
    _write_frontend_log(entry)
    return {"status": "ok"}


# ── Health endpoint ──

CAMERA_HLS_UPSTREAM = "http://10.100.0.202:8888"
_hls_client: httpx.AsyncClient | None = None

async def _get_hls_client():
    global _hls_client
    if _hls_client is None or _hls_client.is_closed:
        _hls_client = httpx.AsyncClient(timeout=15.0)
    return _hls_client

@app.get("/api/camera-hls/{path:path}")
async def camera_hls_proxy(path: str, request: Request):
    client = await _get_hls_client()
    url = f"{CAMERA_HLS_UPSTREAM}/{path}"
    qs = str(request.url.query)
    if qs:
        url = f"{url}?{qs}"
    try:
        resp = await client.get(url)
        headers = {}
        ct = resp.headers.get("content-type")
        if ct:
            headers["content-type"] = ct
        headers["access-control-allow-origin"] = "*"
        headers["cache-control"] = "no-cache"
        return Response(content=resp.content, status_code=resp.status_code, headers=headers)
    except Exception:
        raise HTTPException(status_code=502, detail="Camera HLS upstream unreachable")

@app.get("/api/health")
def api_health():
    with source_lock:
        rtsp_count = len(running_sources)
    with upload_sources_lock:
        upload_count = len(upload_sources)
    return {
        "status": "ok",
        "active_mode": active_mode,
        "rtsp_sources": rtsp_count,
        "uploads": upload_count,
        "inference_paused": _inference_pause_remaining_seconds() > 0,
        "inference_resume_in_seconds": int(math.ceil(_inference_pause_remaining_seconds())),
    }


# ── Overlay endpoints ──

@app.get("/api/overlays")
def get_overlays():
    with overlay_lock:
        try:
            return {k: dict(v) for k, v in overlays.items()}
        except:
            return dict(overlays)

@app.get("/api/overlays/{name}")
def get_overlay(name: str):
    ensure_overlay(name)
    with overlay_lock:
        try:
            return dict(overlays[name])
        except:
            return {"heatmap": True, "heatmap_full": True, "heatmap_trails": True, "trails": True, "bboxes": True}

@app.post("/api/overlays/{name}")
def set_overlay(name: str, update: OverlayUpdate):
    ensure_overlay(name)
    with overlay_lock:
        try:
            cur = dict(overlays[name])
        except:
            cur = {
                "heatmap": True,
                "heatmap_full": True,
                "heatmap_trails": True,
                "trails": True,
                "bboxes": True,
                "confidence": 0.15,
            }

        new_confidence = cur.get("confidence", 0.15)
        if update.confidence is not None:
            new_confidence = max(0.05, min(0.95, update.confidence))

        heatmap_val = update.heatmap if update.heatmap is not None else cur.get("heatmap", True)
        new_state = {
            "heatmap": heatmap_val,
            "heatmap_full": update.heatmap_full if update.heatmap_full is not None else cur.get("heatmap_full", heatmap_val),
            "heatmap_trails": update.heatmap_trails if update.heatmap_trails is not None else cur.get("heatmap_trails", heatmap_val),
            "trails": update.trails if update.trails is not None else cur.get("trails", True),
            "bboxes": update.bboxes if update.bboxes is not None else cur.get("bboxes", True),
            "confidence": new_confidence,
        }
        # Preserve mode context and label behavior set by mode configs.
        if "active_mode" in cur:
            new_state["active_mode"] = cur.get("active_mode")
        if "bbox_label" in cur:
            new_state["bbox_label"] = cur.get("bbox_label")
        # Keep heatmap children coherent.
        if not new_state["heatmap"]:
            new_state["heatmap_full"] = False
            new_state["heatmap_trails"] = False
        else:
            # Turning heatmap back on should re-enable per-vehicle heat layer by default
            # unless caller explicitly controls heatmap_trails.
            if update.heatmap is True and update.heatmap_trails is None:
                new_state["heatmap_trails"] = True
            if update.heatmap_full:
                new_state["heatmap"] = True
                if update.heatmap_trails is None:
                    new_state["heatmap_trails"] = False
            if update.heatmap_full is False and update.heatmap_trails is None:
                new_state["heatmap_trails"] = True
            if update.heatmap_trails:
                new_state["heatmap"] = True
        if cur.get("active_mode") == "congestion":
            # Congestion mode: only heatmaps, no trails/boxes.
            new_state["trails"] = False
            new_state["bboxes"] = False
        overlays[name] = dict(new_state)
        print(
            f"[OVERLAY] {name} updated: heatmap={new_state['heatmap']}, "
            f"heatmap_full={new_state['heatmap_full']}, heatmap_trails={new_state['heatmap_trails']}, "
            f"trails={new_state['trails']}, bboxes={new_state['bboxes']}, confidence={new_state['confidence']}"
        )

    cfg = load_rtsp_config()
    cfg.setdefault("overlays", {})[name] = new_state
    save_rtsp_config(cfg)
    return new_state


# ── Confidence endpoints ──

@app.get("/api/confidence/{name}")
def get_confidence(name: str):
    ensure_overlay(name)
    with overlay_lock:
        try:
            conf = overlays[name].get("confidence", 0.15)
        except:
            conf = 0.15
    return {"source": name, "confidence": conf}

@app.post("/api/confidence/{name}")
def set_confidence(name: str, update: ConfidenceUpdate):
    ensure_overlay(name)
    new_conf = max(0.05, min(0.95, update.confidence))

    with overlay_lock:
        try:
            cur = dict(overlays[name])
        except:
            cur = {
                "heatmap": True,
                "heatmap_full": True,
                "heatmap_trails": True,
                "trails": True,
                "bboxes": True,
                "confidence": 0.15,
            }
        cur["confidence"] = new_conf
        overlays[name] = cur

    print(f"[CONFIDENCE] {name} set to {new_conf}")

    cfg = load_rtsp_config()
    cfg.setdefault("overlays", {})[name] = cur
    save_rtsp_config(cfg)
    return {"source": name, "confidence": new_conf}

@app.get("/api/mode/confidence")
def get_mode_confidence():
    return {"mode_confidence": MODE_CONFIDENCE, "active_mode": active_mode}


# ── Source endpoints ──

@app.get("/api/sources")
def list_sources():
    cfg = load_rtsp_config()
    active = set(running_sources.keys())
    out = []

    for i, url in enumerate(cfg.get("rtsp_links", [])):
        idx = i + 1
        name = source_display_name(url)
        ensure_overlay(name)
        out.append({
            "index": idx,
            "name": name,
            "url": mask_rtsp(url),
            "active": idx in active,
            "start_time": running_sources[idx].get("start_time") if idx in active else None,
        })

    return {"sources": out, "active_sources": sorted(active)}

@app.post("/api/sources/start")
def start_source(req: SourceStartRequest):
    global active_mode
    _ensure_inference_can_start()

    cfg = load_rtsp_config()
    rtsp_links = cfg.get("rtsp_links", [])
    if req.index < 1 or req.index > len(rtsp_links):
        raise HTTPException(400, "Invalid source index or empty rtsp_links config")

    requested_mode = normalize_mode(req.mode)
    source_mode = requested_mode or normalize_mode(active_mode)
    if not source_mode:
        raise HTTPException(400, "No route selected. Open a route in UI before starting inference.")
    if requested_mode and active_mode and requested_mode != active_mode:
        print(f"[MODE] Running mixed modes in parallel: global={active_mode}, source={requested_mode}")
    if requested_mode and not active_mode:
        active_mode = requested_mode

    url = rtsp_links[req.index - 1]
    name = source_display_name(url)

    base_overlay = route_mode_overlay(source_mode)
    mode_confidence = route_mode_confidence(source_mode)
    overlay_config = {**base_overlay, "confidence": mode_confidence, "active_mode": source_mode}

    with source_lock:
        if req.index in running_sources:
            print(f"[START] Source {req.index} ({name}) already active, skipping restart but updating config.")
            # Update the shared overlay dict so the running process picks up changes
            with overlay_lock:
                overlays[name] = dict(overlay_config)
            
            if PERSIST_ACTIVE_SOURCES:
                # Ensure it is in active_sources config
                active = cfg.get("active_sources") or []
                if req.index not in active:
                    active.append(req.index)
                cfg["active_sources"] = active
                save_rtsp_config(cfg)
                
            return {"status": "started", "index": req.index, "mode": source_mode, "info": "already_running"}

        # Calculate total active streams (including this new one) for dynamic GPU allocation
        with upload_sources_lock:
            upload_count = len(upload_sources)
        active_streams = len(running_sources) + upload_count + 1

        if len(running_sources) >= MAX_RTSP_STREAMS:
            raise HTTPException(429, f"Maximum RTSP streams reached ({MAX_RTSP_STREAMS}).")

        process, stop = start_source_callback(req.index, url, name, overlay_config, active_streams)
        if process is None:
            raise HTTPException(500, "Failed to start source")
        running_sources[req.index] = {"process": process, "stop": stop, "start_time": time.time(), "name": name}

    with overlay_lock:
        overlays[name] = dict(overlay_config)
    print(f"[OVERLAY] {name} set from mode config: {overlay_config}")

    _ffmpeg_monitor_sources.add(name)
    ffmpeg_next_start[name] = time.time() + 0.15
    _start_ffmpeg_publisher_async(name)
    _ensure_ffmpeg_monitor()

    if PERSIST_ACTIVE_SOURCES:
        active = cfg.get("active_sources") or []
        if req.index not in active:
            active.append(req.index)
        cfg["active_sources"] = active
        save_rtsp_config(cfg)
    return {"status": "started", "index": req.index, "mode": source_mode}

@app.post("/api/sources/stop")
def stop_source(req: SourceIndexRequest):
    cfg = load_rtsp_config()
    name = None
    if 1 <= req.index <= len(cfg.get("rtsp_links", [])):
        name = source_display_name(cfg["rtsp_links"][req.index - 1])

    with source_lock:
        src = running_sources.pop(req.index, None)
        if src:
            src["stop"].set()
            # Aggressively terminate the process
            proc = src.get("process")
            if proc:
                try:
                    proc.terminate()
                    # Wait briefly for graceful exit, then force kill if needed
                    import time
                    t_start = time.time()
                    while proc.is_alive() and time.time() - t_start < 1.0:
                        time.sleep(0.05)
                    if proc.is_alive():
                        proc.kill()
                except:
                    pass

    if name:
        _ffmpeg_monitor_sources.discard(name)
        _stop_ffmpeg_publisher(name)
        # Clear stale UI state
        with metrics_lock:
            metrics.pop(name, None)
        with overlay_lock:
            overlays.pop(name, None)
        with frame_lock:
            frame_buffer.pop(name, None)
            frame_bgr_buffer.pop(name, None)
            frame_sequences.pop(name, 0)
        with raw_frame_lock:
            raw_frame_buffer.pop(name, None)
            raw_frame_bgr_buffer.pop(name, None)
            raw_frame_sequences.pop(name, None)

    if PERSIST_ACTIVE_SOURCES:
        cfg["active_sources"] = [i for i in cfg.get("active_sources", []) if i != req.index]
        save_rtsp_config(cfg)
    
    # Flush GPU memory if all sources stopped
    if not running_sources:
        try:
            import torch
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
        except:
            pass

    return {"status": "stopped", "index": req.index}

@app.post("/api/sources/stop_all")
def stop_all_sources():
    global active_mode
    count = _stop_all_sources()
    active_mode = None
    return {"status": "stopped", "count": count}


# ── Metrics / Alerts endpoints ──

@app.get("/api/metrics")
def get_metrics():
    return get_all_metrics()

@app.get("/api/processed/{name}/ready")
def processed_ready(name: str):
    proc = ffmpeg_publishers.get(name)
    running = proc is not None and proc.poll() is None
    has_frames = ffmpeg_frame_sequences.get(name, 0) > 0
    if not (running and has_frames):
        print(f"[READY] {name}: running={running} has_frames={has_frames} proc={proc is not None} poll={proc.poll() if proc else 'N/A'} seq={ffmpeg_frame_sequences.get(name, 0)}")
        return {"name": name, "ready": False}

    # Avoid WHEP 404 races: only report ready once MediaMTX has materialized
    # and marked processed_<name> as ready.
    path_name = f"processed_{name}"
    try:
        r = http_requests.get(
            f"{MEDIAMTX_API}/v3/paths/get/{path_name}",
            timeout=0.8,
        )
        if r.status_code != 200:
            return {"name": name, "ready": False}
        data = r.json() if r.content else {}
        return {"name": name, "ready": bool(data.get("ready"))}
    except Exception:
        return {"name": name, "ready": False}

@app.get("/api/alerts")
def api_get_alerts(limit: int = 20):
    return {"alerts": get_alerts(limit)}

@app.get("/api/alerts/{alert_id}/screenshot")
def api_get_alert_screenshot(alert_id: str):
    screenshot_path = ALERTS_DIR / f"{alert_id}.jpg"
    if not screenshot_path.exists():
        raise HTTPException(404, "Screenshot not found")
    return FileResponse(screenshot_path, media_type="image/jpeg")

@app.delete("/api/alerts")
def api_clear_alerts():
    clear_alerts()
    return {"status": "cleared"}


# ── Report generation endpoint ──

@app.post("/api/report/{name}")
def api_generate_report(name: str):
    """Generate a Markdown incident report for a given source.

    Captures current frame screenshot, metrics, and alerts.
    Returns a ZIP file containing the report and screenshots.
    """
    import zipfile
    import io

    print(f"[REPORT] Generating report for {name}")
    
    # Grab current frame screenshot
    with frame_lock:
        screenshot = frame_buffer.get(name)
        if screenshot and not isinstance(screenshot, (bytes, bytearray)):
            print(f"[REPORT] WARNING: Screenshot for {name} is {type(screenshot)}, ignoring")
            screenshot = None
        
        if screenshot:
            print(f"[REPORT] Found screenshot for {name} ({len(screenshot)} bytes)")
        else:
            print(f"[REPORT] WARNING: No screenshot found for {name}")

    # Grab current metrics for this source
    with metrics_lock:
        source_metrics = metrics.get(name) or {}
        # Inject start_time from running_sources or upload_sources
        source_start = None
        
        # Check RTSP sources
        with source_lock:
            for info in running_sources.values():
                if info.get("name") == name:
                    source_start = info.get("start_time")
                    break
        
        # Check Uploads if not found
        if not source_start:
             with upload_sources_lock:
                info = upload_sources.get(name)
                if info:
                    source_start = info.get("start_time")
        
        if source_start:
            source_metrics["start_time"] = source_start

        if source_metrics:
            print(f"[REPORT] Found metrics for {name}")
        else:
            print(f"[REPORT] WARNING: No metrics found for {name}")

    # Resolve per-source mode (overrides global active_mode)
    report_mode = _resolve_source_mode(name, source_metrics)

    # Forensics special: Inject data from sam_results
    if report_mode == "forensics":
        try:
            from sam import sam_results, sam_results_lock
            with sam_results_lock:
                sam_info = sam_results.get(name)
                if sam_info:
                    source_metrics["prompt"] = sam_info.get("prompt")
                    source_metrics["detection_count"] = sam_info.get("count", 0)
                    source_metrics["session_history"] = sam_info.get("session_history", [])
        except Exception as e:
            print(f"[REPORT] Failed to inject SAM metrics: {e}")

    # Grab alerts filtered for this source
    with alerts_lock:
        source_alerts = [a for a in alerts if a.get("source") == name]
        print(f"[REPORT] Found {len(source_alerts)} alerts for {name}")

    # Forensics special: Trigger VLM Analysis
    vlm_narrative = None
    if report_mode == "forensics":
        # Check if SAM has actually analyzed anything
        with sam_results_lock:
            sam_info = sam_results.get(name)
            if not sam_info:
                raise HTTPException(400, "Forensic data matching this feed not found. Did you start a forensic search?")

            history = sam_info.get("session_history", [])
            if not history:
                raise HTTPException(400, "Forensic analysis cycle in progress. Please wait for at least one detection or analysis frame to be recorded.")

        try:
            from sam import generate_vlm_analysis
            vlm_narrative = generate_vlm_analysis(name)
        except Exception as e:
            print(f"[REPORT] VLM Analysis failed: {e}")

    # For all upload streams: pull Gemini forensics job data to power comprehensive report.
    if not vlm_narrative:
        upload_forensics_job_id = None
        upload_original_name = ""
        upload_file_path = ""
        with upload_sources_lock:
            upload_info = upload_sources.get(name)
            if upload_info:
                upload_forensics_job_id = upload_info.get("forensics_job_id")
                upload_original_name = str(upload_info.get("original_name") or "")
                upload_file_path = str(upload_info.get("file_path") or "")
        # Guard against stale per-upload mapping (e.g., reused upload slots like upload1).
        # For reports, prefer the latest DONE matching Gemini job for this file name.
        latest_matching_job_id = None
        try:
            latest_matching_job_id = (
                forensics_events.latest_job_for_video(upload_original_name, require_done=True)
                or forensics_events.latest_job_for_video(Path(upload_file_path).name, require_done=True)
            )
        except Exception:
            latest_matching_job_id = None
        if latest_matching_job_id and latest_matching_job_id != upload_forensics_job_id:
            print(
                f"[REPORT] Using latest matching forensics job for {name}: "
                f"{latest_matching_job_id} (was {upload_forensics_job_id})"
            )
            upload_forensics_job_id = latest_matching_job_id
        if upload_forensics_job_id:
            try:
                f_results = forensics_events.get_results(upload_forensics_job_id)
                if f_results and not f_results.get("error"):
                    master_summary = f_results.get("master_summary", "")
                    events = f_results.get("events", [])
                    source_metrics["forensics_event_count"] = f_results.get("event_count", 0)
                    source_metrics["forensics_events"] = events[:20]
                    parts = []
                    if master_summary:
                        parts.append(master_summary)
                    if events:
                        parts.append("\nDetected Events:")
                        for ev in events[:12]:
                            ts = ev.get("timestamp_str") or str(ev.get("timestamp", ""))
                            desc = ev.get("description", "")
                            if desc:
                                parts.append(f"  [{ts}] {desc}" if ts else f"  {desc}")
                    if parts:
                        vlm_narrative = "\n".join(parts)
                    print(f"[REPORT] Gemini forensics data injected: {source_metrics.get('forensics_event_count', 0)} events")
                else:
                    # Job still running — include live brief so report shows progress
                    f_status = forensics_events.get_status(upload_forensics_job_id)
                    if f_status:
                        pct = f_status.get("progress_percent", 0)
                        brief = f_status.get("live_brief", "")
                        if brief:
                            vlm_narrative = f"[IRIS Drone Analytics in progress — {pct:.0f}% complete]\n\n{brief}"
                            print(f"[REPORT] Analytics job {upload_forensics_job_id} still running ({pct:.0f}%)")
            except Exception as e:
                print(f"[REPORT] Failed to pull Gemini forensics data for {name}: {e}")

    try:
        pdf_bytes = generate_report(
            source_name=name,
            screenshot_bytes=screenshot,
            metrics_data=source_metrics,
            alerts_list=source_alerts,
            active_mode=report_mode,
            vlm_narrative=vlm_narrative,
        )
        # print(f"[REPORT] PDF generated successfully ({len(pdf_bytes)} bytes)")
    except Exception as e:
        print(f"[REPORT] ERROR generating PDF: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Report generation failed: {e}")

    # Return PDF directly
    if not isinstance(pdf_bytes, (bytes, bytearray)):
        print(f"[REPORT] CRITICAL: generate_report returned {type(pdf_bytes)}, forcing cleanup")
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin-1')
            
    pdf_buffer = io.BytesIO(pdf_bytes)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    pdf_filename = f"iris_drone_analytics_{name}_{timestamp}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{pdf_filename}"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


# ── Stream endpoint ──

@app.get("/api/stream/{name}")
def stream_video(name: str):
    def generate():
        last_processed_seq = -1
        last_raw_seq = -1
        latched_processed_frame = None
        while True:
            frame = None

            with frame_lock:
                processed_seq = frame_sequences.get(name, 0)
                if processed_seq > last_processed_seq:
                    frame = frame_buffer.get(name)
                    if frame is not None:
                        latched_processed_frame = frame
                    last_processed_seq = processed_seq

            # Once processed frames appear, keep serving processed output.
            # This avoids raw frames overwhelming VCC overlays when processed cadence is lower.
            if latched_processed_frame is not None:
                frame = latched_processed_frame

            if frame is None:
                with raw_frame_lock:
                    raw_seq = raw_frame_sequences.get(name, 0)
                    if raw_seq > last_raw_seq:
                        frame = raw_frame_buffer.get(name)
                        last_raw_seq = raw_seq

            if frame is not None:
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
            else:
                # No fresh frame on either path yet.
                time.sleep(0.01)

    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "X-Accel-Buffering": "no",
        },
    )

@app.get("/api/raw/{name}")
def stream_raw_video(name: str):
    def generate():
        last_seq = -1
        while True:
            with raw_frame_lock:
                while raw_frame_sequences.get(name, 0) <= last_seq:
                    if not raw_frame_lock.wait(timeout=0.2):
                        break

                frame = raw_frame_buffer.get(name)
                last_seq = raw_frame_sequences.get(name, 0)

            if frame:
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
            time.sleep(0.001)

    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "X-Accel-Buffering": "no",
        },
    )


# ── Jobs endpoints ──

@app.get("/api/jobs")
def list_jobs():
    with jobs_lock:
        return {"jobs": list(jobs.values())}

@app.get("/api/jobs/{job_id}")
def get_job(job_id: str):
    with jobs_lock:
        if job_id in jobs:
            return jobs[job_id]
    raise HTTPException(404, "Job not found")


# ── Gemini Forensics endpoints ──

@app.get("/api/forensics/health")
def api_forensics_health():
    return forensics_events.health()

@app.post("/api/forensics/upload")
async def api_forensics_upload(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(400, "Empty filename")
    payload = await file.read()
    if not payload:
        raise HTTPException(400, "No file content")
    job_id = forensics_events.create_job(file.filename, payload)
    return JSONResponse({"job_id": job_id}, status_code=202)

@app.get("/api/forensics/job/{job_id}/status")
def api_forensics_status(job_id: str):
    data = forensics_events.get_status(job_id)
    if not data:
        raise HTTPException(404, "Job not found")
    return data

@app.get("/api/forensics/job/{job_id}/results")
def api_forensics_results(job_id: str):
    data = forensics_events.get_results(job_id)
    if not data:
        raise HTTPException(404, "Job not found")
    if data.get("error"):
        raise HTTPException(425, data["error"])
    return data

@app.get("/api/forensics/job/{job_id}/frame/{n}")
def api_forensics_frame(
    job_id: str,
    n: int,
    event_id: Optional[int] = None,
):
    frames_dir = forensics_events.get_frames_dir(job_id)
    frame_path = None
    used_event_frame = False
    if frames_dir:
        # 1. Materialized event image (VCC-processed snapshot, highest priority).
        if event_id is not None:
            event_path = frames_dir / f"event_{int(event_id):05d}.jpg"
            if event_path.exists():
                frame_path = event_path
                used_event_frame = True

        # 2. Exact processed frame (Supervision-annotated by live VCC pipeline).
        if frame_path is None:
            p = frames_dir / f"processed_{n:05d}.jpg"
            if p.exists():
                frame_path = p

        # 3. Search ±5 nearby processed frames (Gemini may report a slightly off index).
        if frame_path is None:
            for delta in range(1, 6):
                for sign in (-1, 1):
                    idx = n + sign * delta
                    if idx < 0:
                        continue
                    cand = frames_dir / f"processed_{idx:05d}.jpg"
                    if cand.exists():
                        frame_path = cand
                        break
                if frame_path is not None:
                    break

        # 4. Raw frame — only last resort (never shown as event image if processed exists).
        if frame_path is None:
            r = frames_dir / f"frame_{n:05d}.jpg"
            if r.exists():
                frame_path = r

    if frame_path is None:
        frame_path = forensics_events.get_frame_path(job_id, n)
    if not frame_path:
        raise HTTPException(404, "Frame not found")
    cache_header = "public, max-age=86400"
    if used_event_frame:
        cache_header = "public, max-age=604800, immutable"
    return FileResponse(frame_path, media_type="image/jpeg", headers={"Cache-Control": cache_header})

@app.get("/api/forensics/job/{job_id}/pdf")
def api_forensics_pdf(job_id: str, unified: bool = False):
    data = forensics_events.get_pdf(job_id, force_unified=unified)
    if not data:
        raise HTTPException(404, "Job not complete or not found")
    pdf_bytes, name = data
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{name}"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

@app.get("/api/forensics/jobs/latest")
def api_forensics_latest():
    return forensics_events.latest_job()


@app.post("/api/forensics/job/{job_id}/cancel")
def api_forensics_cancel_job(job_id: str):
    ok = forensics_events.cancel_job(job_id, reason="Cancelled by user")
    if not ok:
        raise HTTPException(404, "Job not found or already in terminal state")
    return {"status": "cancelled", "job_id": job_id}


@app.post("/api/forensics/cancel_all")
def api_forensics_cancel_all():
    cancelled = forensics_events.cancel_active_jobs(reason="Cancelled by user (stop-all)")
    return {"cancelled": cancelled}


# ── Realtime Forensics endpoints ──

@app.post("/api/realtime-forensics/start")
def api_rt_forensics_start():
    return rt_forensics.start_session()

@app.post("/api/realtime-forensics/stop")
def api_rt_forensics_stop():
    return rt_forensics.stop_session()

@app.get("/api/realtime-forensics/status")
def api_rt_forensics_status():
    return rt_forensics.get_status()

@app.get("/api/realtime-forensics/events")
def api_rt_forensics_events():
    return {"events": rt_forensics.get_all_events()}

@app.get("/api/realtime-forensics/event/{event_id}/frame")
def api_rt_forensics_event_frame(event_id: int):
    fpath = rt_forensics.get_event_frame(event_id)
    if not fpath or not Path(fpath).exists():
        raise HTTPException(404, "Event frame not found")
    return FileResponse(fpath, media_type="image/jpeg")

@app.get("/api/realtime-forensics/event/{event_id}/clip")
def api_rt_forensics_event_clip(event_id: int):
    fpath = rt_forensics.get_event_clip(event_id)
    if not fpath or not Path(fpath).exists():
        raise HTTPException(404, "Clip not ready or not found")
    return FileResponse(fpath, media_type="video/mp4", filename=f"alert_{event_id}.mp4")

@app.get("/api/realtime-forensics/event/{event_id}/clip/status")
def api_rt_forensics_event_clip_status(event_id: int):
    return rt_forensics.get_event_clip_status(event_id)

@app.get("/api/realtime-forensics/report/pdf")
def api_rt_forensics_report_pdf():
    path = rt_forensics.generate_report_pdf()
    if not path:
        raise HTTPException(status_code=404, detail="No session data for report")
    return FileResponse(path, media_type="application/pdf", filename="iris_realtime_forensics_report.pdf")


# ── Crowd Report endpoints ──

_crowd_report_sessions: Dict[str, dict] = {}
_crowd_report_sessions_lock = threading.Lock()

@app.post("/api/crowd-report/init")
def api_crowd_report_init(req: UploadInitRequest):
    filename = Path(req.filename or "upload.mp4").name
    _validate_upload_format(filename)
    if req.size <= 0:
        raise HTTPException(400, "Invalid upload size")

    requested_chunk = int(req.chunk_size or DEFAULT_UPLOAD_CHUNK_SIZE)
    chunk_size = min(max(1 * 1024 * 1024, requested_chunk), MAX_UPLOAD_CHUNK_SIZE)
    total_chunks = (req.size + chunk_size - 1) // chunk_size

    upload_id = uuid.uuid4().hex
    target = crowd_report.CROWD_REPORT_UPLOAD_DIR / f"{upload_id}_{filename}"
    crowd_report.CROWD_REPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    target.touch(exist_ok=False)

    with _crowd_report_sessions_lock:
        _crowd_report_sessions[upload_id] = {
            "file_path": str(target),
            "filename": filename,
            "size": int(req.size),
            "chunk_size": int(chunk_size),
            "total_chunks": int(total_chunks),
            "received": {},
            "updated_at": time.time(),
        }

    return {"upload_id": upload_id, "chunk_size": int(chunk_size), "total_chunks": int(total_chunks)}

@app.post("/api/crowd-report/chunk")
async def api_crowd_report_chunk(request: Request, upload_id: str, index: int):
    if index < 0:
        raise HTTPException(400, "Invalid chunk index")
    with _crowd_report_sessions_lock:
        info = _crowd_report_sessions.get(upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        chunk_size = int(info["chunk_size"])
        total_chunks = int(info["total_chunks"])
        file_path = str(info["file_path"])
    if index >= total_chunks:
        raise HTTPException(400, "Chunk index out of range")

    body = await request.body()
    if not body:
        raise HTTPException(400, "Empty chunk")

    offset = index * chunk_size
    with open(file_path, "r+b") as f:
        f.seek(offset)
        f.write(body)

    with _crowd_report_sessions_lock:
        info = _crowd_report_sessions.get(upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        received = info.setdefault("received", {})
        received[index] = len(body)
        info["updated_at"] = time.time()
        done = len(received) >= int(info["total_chunks"])

    return {"ok": True, "done": done}

@app.post("/api/crowd-report/complete")
def api_crowd_report_complete(req: UploadCompleteRequest):
    with _crowd_report_sessions_lock:
        info = _crowd_report_sessions.get(req.upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        file_path = Path(str(info["file_path"]))
        filename = str(info["filename"])
        total_chunks = int(info["total_chunks"])
        expected_size = int(info["size"])
        received = dict(info.get("received", {}))
        _crowd_report_sessions.pop(req.upload_id, None)

    if len(received) != total_chunks:
        raise HTTPException(400, "Upload incomplete")
    if sum(int(v) for v in received.values()) != expected_size:
        raise HTTPException(400, "Upload size mismatch")
    if not file_path.exists():
        raise HTTPException(400, "Uploaded file missing")

    job_id = crowd_report.create_job_from_path(filename, str(file_path))
    return JSONResponse({"job_id": job_id}, status_code=202)

@app.get("/api/crowd-report/job/{job_id}/stream")
def api_crowd_report_stream(job_id: str):
    """MJPEG stream of live inference frames during report generation."""
    def generate():
        last_seq = -1
        idle_count = 0
        while True:
            frame_data, seq = crowd_report.get_live_frame(job_id)
            if frame_data and seq > last_seq:
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_data + b'\r\n')
                last_seq = seq
                idle_count = 0
            else:
                idle_count += 1
                # Stop if no frames for ~10 seconds (job likely finished)
                if idle_count > 200:
                    break
            time.sleep(0.05)

    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
    )

@app.get("/api/crowd-report/job/{job_id}/status")
def api_crowd_report_status(job_id: str):
    data = crowd_report.get_status(job_id)
    if not data:
        raise HTTPException(404, "Job not found")
    return data

@app.get("/api/crowd-report/job/{job_id}/pdf")
def api_crowd_report_pdf(job_id: str):
    pdf_path = crowd_report.get_pdf_path(job_id)
    if not pdf_path or not os.path.exists(pdf_path):
        raise HTTPException(404, "PDF not ready or job not found")
    return FileResponse(pdf_path, media_type="application/pdf", filename="IRIS_Crowd_Report.pdf")

@app.get("/api/crowd-report/jobs/latest")
def api_crowd_report_latest():
    data = crowd_report.latest_job()
    if not data:
        return {"id": None}
    return data


# ── Live Crowd Monitoring endpoints ──

@app.post("/api/crowd-live/start")
async def api_crowd_live_start(request: Request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    cameras = body.get("cameras") if body else None
    return rt_crowd.start_session(selected_cameras=cameras)

@app.post("/api/crowd-live/stop")
def api_crowd_live_stop():
    return rt_crowd.stop_session()

@app.get("/api/crowd-live/cameras")
def api_crowd_live_cameras():
    """Return list of available cameras for selection."""
    return rt_crowd.get_available_cameras()

@app.post("/api/crowd-live/cameras/add")
async def api_crowd_live_cameras_add(request: Request):
    """Add a camera to crowd analysis. Auto-starts session if not running."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    name = body.get("name", "")
    url = body.get("url", "")
    return rt_crowd.add_camera(name, url)

@app.post("/api/crowd-live/cameras/remove")
async def api_crowd_live_cameras_remove(request: Request):
    """Remove a camera from crowd analysis."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    name = body.get("name", "")
    return rt_crowd.remove_camera(name)

@app.get("/api/crowd-live/groups")
def api_crowd_live_groups():
    return rt_crowd.get_camera_groups()

@app.post("/api/crowd-live/groups/create")
async def api_crowd_live_groups_create(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    name = body.get("name", "")
    camera_names = body.get("camera_names")
    return rt_crowd.create_camera_group(name, camera_names)

@app.post("/api/crowd-live/groups/delete")
async def api_crowd_live_groups_delete(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    return rt_crowd.delete_camera_group(body.get("name", ""))

@app.post("/api/crowd-live/groups/activate")
async def api_crowd_live_groups_activate(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    return rt_crowd.activate_camera_group(body.get("name", ""))

@app.get("/api/crowd-live/status")
def api_crowd_live_status():
    return rt_crowd.get_status()

@app.get("/api/crowd-live/events")
def api_crowd_live_events():
    return {"events": rt_crowd.get_all_events()}


# ── Magicboxhub camera-tree integration ──────────────────────────────────
# Lets the operator pick cameras directly from app.magicboxhub.net's
# station/device/camera tree and bulk-add them to the crowd-live session.

@app.get("/api/crowd-live/magicbox/status")
def api_mbx_status():
    return magicboxhub_api.status()


@app.post("/api/crowd-live/magicbox/login")
async def api_mbx_login(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    email = (body or {}).get("email", "")
    password = (body or {}).get("password", "")
    try:
        magicboxhub_api.login(email, password)
    except magicboxhub_api.MagicboxError as exc:
        return {"status": "error", "message": str(exc)}
    return {"status": "ok", **magicboxhub_api.status()}


@app.post("/api/crowd-live/magicbox/logout")
def api_mbx_logout():
    magicboxhub_api.logout()
    return {"status": "ok"}


def _restart_crowd_worker_async() -> bool:
    """SIGKILL any running crowd_worker.py and respawn it.

    crowd_worker.py reads its camera list once at startup via /api/crowd-worker/config,
    so we have to bounce it for newly-added cameras to appear in the dashboard tiles.
    Mirrors the pattern used by /api/crowd-live/cameras/apply.
    """
    import signal, subprocess, threading
    try:
        result = subprocess.run(['pgrep', '-f', 'crowd_worker.py'], capture_output=True, text=True)
        for pid in (result.stdout or "").strip().split('\n'):
            if pid.strip():
                try:
                    os.kill(int(pid.strip()), signal.SIGKILL)
                except (ProcessLookupError, ValueError):
                    pass

        # Drop cached tiles from the previous camera set — otherwise
        # /api/crowd/analysis/latest keeps serving stale entries that
        # render as ghost cameras in the dashboard.
        with _crowd_worker_lock:
            _crowd_worker_analysis.clear()
            _crowd_worker_history.clear()

        def _spawn():
            import time as _t
            _t.sleep(2)
            venv_python = str(Path(__file__).resolve().parent / ".venv" / "bin" / "python")
            worker_script = str(Path(__file__).resolve().parent / "crowd_worker.py")
            subprocess.Popen(
                [venv_python, "-u", worker_script],
                stdout=open('/tmp/crowd_worker.log', 'w'),
                stderr=subprocess.STDOUT,
                cwd=str(Path(__file__).resolve().parent),
                start_new_session=True,
            )
            print("[MBX] crowd_worker restarted")
        threading.Thread(target=_spawn, daemon=True).start()
        return True
    except Exception as exc:
        print(f"[MBX] crowd_worker restart failed: {exc}")
        return False


def _mbx_camera_internal_name(cam: dict) -> str:
    """Stable, dedupable internal name for a magicbox camera."""
    cid = (cam.get("cameraId") or cam.get("id") or "").strip()
    if cid:
        return f"MBX_{cid[:8]}"
    # Fallback to slugified human name (rare)
    name = (cam.get("name") or "camera").strip()
    safe = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
    return f"MBX_{safe[:24] or 'cam'}"


_MBX_UUID_PATH_RE = re.compile(
    r"^(rtsp://[^/]+/)([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})(/?.*)$",
    re.IGNORECASE,
)


def _mbx_fix_rtsp_url(url: str) -> str:
    """Prepend `camera_` to the UUID path on magicbox edge RTSP URLs.

    The upstream camera-tree returns `rtsp://<host>:8554/<uuid>`, but the
    USS streamcontroller running on each edge serves the stream under
    `camera_<uuid>` (verified against the gortsplib DESCRIBE response).
    Without this rewrite, every fetch hits a non-existent path and gets
    400 Bad Request, so no frames ever reach the analytics worker.
    """
    if not url:
        return url
    m = _MBX_UUID_PATH_RE.match(url.strip())
    if not m:
        return url  # not a magicbox-shaped URL, leave alone
    prefix, uuid_part, tail = m.group(1), m.group(2), m.group(3)
    return f"{prefix}camera_{uuid_part}{tail}"


@app.get("/api/crowd-live/magicbox/tree")
def api_mbx_tree():
    """Return the magicboxhub camera tree, normalized for the picker UI.

    Each camera carries an `internalName` we'll use when adding it,
    plus an `alreadyAdded` flag so the UI can render it as a checkbox
    that's already checked + disabled.
    """
    try:
        raw = magicboxhub_api.fetch_tree()
    except magicboxhub_api.MagicboxError as exc:
        return JSONResponse(status_code=502, content={"status": "error", "message": str(exc)})

    existing_internal_names = {
        cam["name"] for cam in rt_crowd.get_available_cameras()
    }

    stations_out = []
    for st in raw.get("stations", []) or []:
        devices_out = []
        for dev in st.get("devices", []) or []:
            cams_out = []
            for cam in dev.get("cameras", []) or []:
                rtsp = _mbx_fix_rtsp_url((cam.get("rtsp") or "").strip())
                if not rtsp:
                    continue  # skip cameras without an RTSP URL
                internal = _mbx_camera_internal_name(cam)
                cams_out.append({
                    "cameraId": cam.get("cameraId") or cam.get("id"),
                    "name": cam.get("name") or internal,
                    "internalName": internal,
                    "status": cam.get("status"),
                    "rtsp": rtsp,
                    "alreadyAdded": internal in existing_internal_names,
                })
            if cams_out:
                devices_out.append({
                    "id": dev.get("id"),
                    "name": dev.get("name"),
                    "ip": dev.get("ip"),
                    "status": dev.get("status"),
                    "location": dev.get("location"),
                    "latitude": dev.get("latitude"),
                    "longitude": dev.get("longitude"),
                    "cameras": cams_out,
                })
        if devices_out:
            stations_out.append({
                "name": st.get("name"),
                "division": st.get("division"),
                "devices": devices_out,
            })

    return {"status": "ok", "stations": stations_out}


@app.post("/api/crowd-live/magicbox/add")
async def api_mbx_add(request: Request):
    """Bulk-add a list of magicbox cameras to the crowd-live session.

    Body: { "cameras": [ { "internalName": "MBX_xxxxxxxx", "rtsp": "rtsp://..." }, ... ] }

    Each entry is funnelled through rt_crowd.add_camera, which auto-starts
    or appends to the running session as needed. Existing cameras are
    treated as a successful skip, not an error.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    requested = (body or {}).get("cameras") or []
    if not isinstance(requested, list) or not requested:
        return {"status": "error", "message": "cameras list required"}

    added, skipped, failed = [], [], []
    for entry in requested:
        if not isinstance(entry, dict):
            continue
        name = (entry.get("internalName") or entry.get("name") or "").strip()
        url = _mbx_fix_rtsp_url((entry.get("rtsp") or entry.get("url") or "").strip())
        if not name or not url:
            failed.append({"name": name, "reason": "missing name or rtsp"})
            continue
        try:
            result = rt_crowd.add_camera(name, url)
        except Exception as exc:
            failed.append({"name": name, "reason": f"add_camera raised: {exc}"})
            continue
        if result.get("status") == "ok":
            added.append(name)
        else:
            msg = str(result.get("message", "")).lower()
            if "already exists" in msg:
                skipped.append(name)
            else:
                failed.append({"name": name, "reason": result.get("message") or "unknown"})

    worker_restarted = False
    if added:
        worker_restarted = _restart_crowd_worker_async()

    return {
        "status": "ok",
        "added": added,
        "skipped": skipped,
        "failed": failed,
        "worker_restarted": worker_restarted,
        "cameras": rt_crowd.get_available_cameras(),
    }


def _cw_read_csv_data(date_str: str = "") -> list:
    """Read crowd-worker CSV and return list of row dicts."""
    import csv
    date_str = date_str or datetime.now().strftime('%Y-%m-%d')
    csv_path = _CW_DATA_DIR / f"crowd_events_{date_str}.csv"
    if not csv_path.exists():
        return []
    try:
        with open(csv_path) as f:
            return list(csv.DictReader(f))
    except Exception:
        return []


def _cw_find_heatmaps_for_camera(device_id: str, limit: int = 4) -> list:
    """Find the latest heatmap image files for a camera from ~/heatmaps/."""
    pattern = f"{device_id}_*.jpg"
    files = sorted(Path(_HEATMAP_DIR).glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    result = []
    for f in files[:limit]:
        # Extract timestamp from filename: camera_id_YYYYMMDD_HHMMSS_microseconds.jpg
        parts = f.stem.replace(device_id + '_', '')
        try:
            ts = datetime.strptime(parts[:15], '%Y%m%d_%H%M%S').strftime('%H:%M:%S')
        except Exception:
            ts = ''
        result.append({'path': str(f), 'timestamp': ts, 'count': 0})
    result.reverse()  # oldest first
    return result


def _cw_build_segments_from_csv(rows: list, segment_minutes: int = 10) -> list:
    """Group CSV rows into time-based segments with per-camera data."""
    if not rows:
        return []
    from collections import defaultdict

    # Parse timestamps and group into segments
    segments = []
    current_rows = []
    seg_start_dt = None

    for row in rows:
        ts_str = row.get('timestamp', '')
        try:
            ts_dt = datetime.fromisoformat(ts_str)
        except Exception:
            continue
        if seg_start_dt is None:
            seg_start_dt = ts_dt
        # Close segment if interval exceeded
        if (ts_dt - seg_start_dt).total_seconds() >= segment_minutes * 60 and current_rows:
            segments.append(_cw_finalize_segment(current_rows, len(segments) + 1, seg_start_dt))
            current_rows = []
            seg_start_dt = ts_dt
        current_rows.append((ts_dt, row))

    # Final segment
    if current_rows:
        segments.append(_cw_finalize_segment(current_rows, len(segments) + 1, seg_start_dt))

    return segments


def _cw_finalize_segment(rows: list, seg_idx: int, seg_start_dt) -> dict:
    """Convert a list of (datetime, row) tuples into a segment record for the report."""
    from collections import defaultdict
    seg_end_dt = rows[-1][0]

    per_camera = defaultdict(lambda: {'counts': [], 'timestamps': [], 'thumbnails': [], 'analysis_rows': []})
    all_counts = []

    for ts_dt, row in rows:
        device_id = row.get('deviceId', '')
        count = int(row.get('peopleCount', 0) or 0)
        cam = per_camera[device_id]
        cam['counts'].append(count)
        cam['timestamps'].append(ts_dt.strftime('%H:%M:%S'))
        cam['analysis_rows'].append(row)
        all_counts.append(count)

    # Build per_camera_data in the format _draw_camera_page expects
    per_camera_data = {}
    analysis = {}
    for device_id, cam in per_camera.items():
        cam_name = device_id.replace('camera_', '').replace('_', '.')
        counts = cam['counts']
        analysis_rows = cam['analysis_rows']

        # Build thumbnails from CSV rows' heatmapImageUrl (each row = one capture)
        # Pick up to 4 evenly spaced rows that have existing heatmap files
        available = []
        for j, row in enumerate(analysis_rows):
            hmap_url = row.get('heatmapImageUrl', '')
            if not hmap_url:
                continue
            # Convert URL path to absolute filesystem path
            filename = hmap_url.split('/')[-1]
            fpath = os.path.join(_HEATMAP_DIR, filename)
            if os.path.exists(fpath):
                yolo_ct = int(row.get('yoloCount', 0) or 0)
                available.append({
                    'path': fpath,
                    'timestamp': cam['timestamps'][j],
                    'count': yolo_ct,
                    'row_idx': j,
                })

        # Pick 4 evenly spaced thumbnails from available frames
        if len(available) > 4:
            step = len(available) / 4
            thumbnails = [available[int(i * step)] for i in range(4)]
        else:
            thumbnails = available

        # Main heatmap: use the latest available frame, with its yoloCount
        latest_heatmap = thumbnails[-1]['path'] if thumbnails else ''
        latest_yolo = thumbnails[-1]['count'] if thumbnails else 0

        per_camera_data[device_id] = {
            'camera_name': cam_name,
            'counts': counts,
            'timestamps': cam['timestamps'],
            'avg_count': round(sum(counts) / len(counts), 1) if counts else 0,
            'peak_count': max(counts) if counts else 0,
            'heatmap_path': latest_heatmap,
            'heatmap_frame_count': latest_yolo,
            'num_cameras': 1,
            'thumbnails': thumbnails,
        }

        # Use latest row for Gemini analysis fields
        latest_row = cam['analysis_rows'][-1]
        analysis = {
            'crowd_movement': latest_row.get('crowd_movement', ''),
            'crowd_density': latest_row.get('densityLevel', 'LOW'),
            'sentiment': latest_row.get('sentiment', 'NEUTRAL'),
            'overall_risk': latest_row.get('overall_risk', 'LOW'),
            'safety_precaution': latest_row.get('safety_precaution', ''),
            'visibility_score': int(latest_row.get('visibility_score', 90) or 90),
            'weapon_detected': latest_row.get('weapon_detected', 'NO'),
            'fight_collision_injury': latest_row.get('fight_collision_injury', 'NO'),
            'wrongful_activity': latest_row.get('wrongful_activity', 'NO'),
            'behavior': latest_row.get('behavior', ''),
        }

    return {
        'segment_index': seg_idx,
        'segment_start': seg_start_dt.strftime('%H:%M:%S'),
        'segment_end': seg_end_dt.strftime('%H:%M:%S'),
        'per_capture_counts': all_counts,
        'per_camera_data': per_camera_data,
        'analysis': analysis,
    }


def _cw_build_report(is_master: bool = False, date_str: str = ""):
    """Build PDF report from crowd-worker CSV data + latest heatmap frames."""
    rows = _cw_read_csv_data(date_str)
    if not rows:
        return None
    started = rows[0].get('timestamp', '') if rows else ''
    segment_min = 60 if is_master else 10
    segments = _cw_build_segments_from_csv(rows, segment_minutes=segment_min)
    if not segments:
        return None
    try:
        from realtime_crowd import _build_pdf_report
        return _build_pdf_report(segments, started, is_master=is_master)
    except Exception as e:
        print(f"[CW-REPORT] Error building PDF: {e}")
        import traceback; traceback.print_exc()
        return None


@app.get("/api/crowd-live/report/pdf")
def api_crowd_live_report_pdf():
    # Try crowd-worker data first, fall back to old pipeline
    path = _cw_build_report(is_master=False)
    if not path:
        path = rt_crowd.generate_report_pdf()
    if not path:
        raise HTTPException(status_code=404, detail="No segment data available for report")
    return FileResponse(path, media_type="application/pdf", filename="IRIS_Crowd_Live_Report.pdf")

@app.get("/api/crowd-live/master-report/pdf")
def api_crowd_live_master_report():
    path = _cw_build_report(is_master=True)
    if not path:
        path = rt_crowd.generate_master_report()
    if not path:
        raise HTTPException(status_code=404, detail="No segment data available for master report")
    return FileResponse(path, media_type="application/pdf", filename="IRIS_Crowd_Master_Report.pdf")

@app.get("/api/crowd-live/csv")
def api_crowd_live_csv():
    # Try crowd-worker CSV first
    cw_path = str(_cw_csv_path)
    if os.path.exists(cw_path):
        return FileResponse(cw_path, media_type="text/csv", filename=f"crowd_events_{datetime.now().strftime('%Y-%m-%d')}.csv")
    path = rt_crowd.get_csv_path()
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="No CSV log available")
    return FileResponse(path, media_type="text/csv", filename=f"crowd_events_{datetime.now().strftime('%Y-%m-%d')}.csv")

@app.get("/api/crowd-live/hourly/dates")
def api_crowd_live_hourly_dates():
    """List dates that have crowd-worker CSV data available."""
    dates = []
    for f in sorted(_CW_DATA_DIR.glob("crowd_events_*.csv"), reverse=True):
        d = f.stem.replace("crowd_events_", "")
        if d:
            dates.append(d)
    if not dates:
        return {"dates": rt_crowd.get_available_dates()}
    return {"dates": dates}

@app.get("/api/crowd-live/hourly/slots")
def api_crowd_live_hourly_slots(date: str = ""):
    """List available hourly slots for a date from CSV data."""
    date_str = date or datetime.now().strftime('%Y-%m-%d')
    rows = _cw_read_csv_data(date_str)
    if not rows:
        return {"slots": rt_crowd.get_hourly_slots(date or None)}
    from collections import defaultdict
    hour_data = defaultdict(lambda: {'counts': [], 'cameras': set()})
    for row in rows:
        try:
            ts = datetime.fromisoformat(row.get('timestamp', ''))
            h = ts.hour
            slot_key = f"{h}-{h+1}"
            hour_data[slot_key]['counts'].append(int(row.get('peopleCount', 0) or 0))
            hour_data[slot_key]['cameras'].add(row.get('deviceId', ''))
        except Exception:
            pass
    slots = []
    for slot_key in sorted(hour_data.keys(), key=lambda s: int(s.split('-')[0])):
        d = hour_data[slot_key]
        counts = d['counts']
        num_cameras = len(d['cameras'])
        # Estimate segments (1 per 10-min window)
        num_segments = max(1, len(counts) // (num_cameras * 2)) if num_cameras else 1
        slots.append({
            'slot': slot_key,
            'segments': num_segments,
            'cameras': num_cameras,
            'avg_count': round(sum(counts) / len(counts), 1) if counts else 0,
            'peak_count': max(counts) if counts else 0,
        })
    return {"slots": slots}

@app.get("/api/crowd-live/hourly/report/{slot}")
def api_crowd_live_hourly_report(slot: str, date: str = ""):
    """Generate hourly report from crowd-worker CSV data."""
    date_str = date or datetime.now().strftime('%Y-%m-%d')
    rows = _cw_read_csv_data(date_str)
    if not rows:
        path = rt_crowd.generate_hourly_report(slot, date or None)
        if not path:
            raise HTTPException(status_code=404, detail=f"No data available for slot {slot}")
        return FileResponse(path, media_type="application/pdf", filename=f"IRIS_Crowd_Hourly_{slot}.pdf")

    # Filter rows to the requested hour slot (e.g. "22-23")
    try:
        start_hour = int(slot.split('-')[0])
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid slot format: {slot}")
    filtered = []
    for row in rows:
        try:
            ts = datetime.fromisoformat(row.get('timestamp', ''))
            if ts.hour == start_hour:
                filtered.append(row)
        except Exception:
            pass
    if not filtered:
        path = rt_crowd.generate_hourly_report(slot, date or None)
        if not path:
            raise HTTPException(status_code=404, detail=f"No data for slot {slot}")
        return FileResponse(path, media_type="application/pdf", filename=f"IRIS_Crowd_Hourly_{slot}.pdf")
    # Build segments from filtered rows (10-min segments within the hour)
    segments = _cw_build_segments_from_csv(filtered, segment_minutes=10)
    started = filtered[0].get('timestamp', '')
    try:
        from realtime_crowd import _build_pdf_report
        path = _build_pdf_report(segments, started, is_master=False)
        if not path:
            raise HTTPException(status_code=500, detail="Failed to build hourly report")
        return FileResponse(path, media_type="application/pdf", filename=f"IRIS_Crowd_Hourly_{date_str}_{slot}.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _build_5min_frame_report(camera_frames: dict, start_dt, end_dt, date_str: str) -> str:
    """Build a simple frame-by-frame PDF for the 5-min window.

    camera_frames: {device_id: [{'timestamp': str, 'yoloCount': int, 'heatmap_path': str}, ...]}
    Each camera gets its own section. Every ~5th second snapshot is shown as a
    thumbnail with timestamp and person count.
    """
    from fpdf import FPDF
    import uuid

    class SimplePDF(FPDF):
        def footer(self):
            self.set_y(-10)
            self.set_font('helvetica', 'I', 7)
            self.set_text_color(150, 150, 150)
            self.cell(0, 5, f'IRIS 5-Min Crowd Report | Page {self.page_no()}', 0, 0, 'C')

    pdf = SimplePDF('L', 'mm', 'A4')  # Landscape for more frames per row
    pdf.set_auto_page_break(auto=True, margin=12)

    # Per camera — no separate cover page
    first_cam = True
    for device_id, frames in camera_frames.items():
        cam_name = device_id.replace('camera_', '').replace('_', '.')
        pdf.add_page()

        # Camera header (first page also serves as report title)
        pdf.set_fill_color(10, 25, 75)
        header_h = 22 if first_cam else 16
        pdf.rect(0, 0, 297, header_h, 'F')
        pdf.set_text_color(255, 255, 255)
        if first_cam:
            pdf.set_font("helvetica", "B", 14)
            pdf.set_xy(0, 2)
            pdf.cell(297, 8, "IRIS 5-MINUTE CROWD SNAPSHOT REPORT", align="C")
            pdf.set_font("helvetica", "", 9)
            pdf.set_xy(0, 11)
            pdf.cell(297, 5, f"{start_dt.strftime('%I:%M:%S %p')} - {end_dt.strftime('%I:%M:%S %p')}  |  {date_str}  |  {len(camera_frames)} cameras", align="C")
            pdf.set_font("helvetica", "B", 10)
            pdf.set_xy(0, 16)
            pdf.cell(297, 5, f"{cam_name}  ({len(frames)} snapshots)", align="C")
            first_cam = False
        else:
            pdf.set_font("helvetica", "B", 12)
            pdf.set_xy(0, 3)
            pdf.cell(297, 10, f"{cam_name}  |  {len(frames)} snapshots  |  {start_dt.strftime('%I:%M:%S %p')} - {end_dt.strftime('%I:%M:%S %p')}", align="C")

        # Grid layout: 4 columns x N rows
        thumb_w = 65
        thumb_h = 37
        cols = 4
        x_start = 8
        y_start = header_h + 4
        gap_x = 4
        gap_y = 14
        col = 0
        row_y = y_start

        for frame_info in frames:
            heatmap_path = frame_info.get('heatmap_path', '')
            ts = frame_info.get('timestamp', '')
            count = frame_info.get('yoloCount', 0)

            # Check if we need a new page
            if row_y + thumb_h + gap_y > 200:
                pdf.add_page()
                # Repeat camera header
                pdf.set_fill_color(10, 25, 75)
                pdf.rect(0, 0, 297, 16, 'F')
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("helvetica", "B", 12)
                pdf.set_xy(0, 3)
                pdf.cell(297, 10, f"{cam_name} (cont.)", align="C")
                row_y = y_start
                col = 0

            tx = x_start + col * (thumb_w + gap_x)

            # Draw frame image
            if heatmap_path and os.path.exists(heatmap_path):
                pdf.image(heatmap_path, x=tx, y=row_y, w=thumb_w, h=thumb_h)
            else:
                # Placeholder box
                pdf.set_draw_color(60, 60, 60)
                pdf.rect(tx, row_y, thumb_w, thumb_h)
                pdf.set_xy(tx, row_y + thumb_h / 2 - 3)
                pdf.set_font("helvetica", "", 7)
                pdf.set_text_color(100, 100, 100)
                pdf.cell(thumb_w, 6, "Frame not available", align="C")

            # Timestamp label
            pdf.set_xy(tx, row_y + thumb_h + 0.5)
            pdf.set_font("helvetica", "", 7)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(thumb_w / 2, 4, ts, align="L")

            # Person count label
            pdf.set_xy(tx + thumb_w / 2, row_y + thumb_h + 0.5)
            pdf.set_fill_color(10, 25, 75)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("helvetica", "B", 7)
            pdf.cell(thumb_w / 2, 4, f"{count} persons", align="R", fill=True)

            col += 1
            if col >= cols:
                col = 0
                row_y += thumb_h + gap_y

    report_id = uuid.uuid4().hex[:8]
    reports_dir = Path(__file__).resolve().parent / "data" / "realtime_crowd" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = str(reports_dir / f"crowd_5min_{date_str}_{start_dt.strftime('%H%M%S')}_{report_id}.pdf")
    pdf.output(pdf_path)
    return pdf_path


@app.get("/api/crowd-live/5min/report")
def api_crowd_live_5min_report(start: str = "", period: str = "", date: str = ""):
    """Generate a 5-minute frame-by-frame crowd report.

    Shows every ~5th second snapshot per camera: heatmap frame, timestamp, person count.
    No AI analytics — just raw frames with bounding box counts.
    """
    date_str = date or datetime.now().strftime('%Y-%m-%d')
    if not start:
        raise HTTPException(status_code=400, detail="Missing 'start' parameter (HH:MM:SS)")

    # Parse start time
    try:
        parts = start.split(':')
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        second = int(parts[2]) if len(parts) > 2 else 0
        if period.upper() == 'PM' and hour != 12:
            hour += 12
        elif period.upper() == 'AM' and hour == 12:
            hour = 0
        start_dt = datetime.strptime(f"{date_str} {hour:02d}:{minute:02d}:{second:02d}", "%Y-%m-%d %H:%M:%S")
        from datetime import timedelta as _td
        end_dt = start_dt + _td(minutes=5)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid start time: {start} {period} — {e}")

    rows = _cw_read_csv_data(date_str)
    if not rows:
        raise HTTPException(status_code=404, detail=f"No data for date {date_str}")

    # Filter rows within the 5-min window, group by camera
    from collections import defaultdict
    camera_frames = defaultdict(list)
    for row in rows:
        try:
            ts = datetime.fromisoformat(row.get('timestamp', ''))
            if start_dt <= ts < end_dt:
                device_id = row.get('deviceId', '')
                hmap_url = row.get('heatmapImageUrl', '')
                filename = hmap_url.split('/')[-1] if hmap_url else ''
                heatmap_path = os.path.join(_HEATMAP_DIR, filename) if filename else ''
                camera_frames[device_id].append({
                    'timestamp': ts.strftime('%H:%M:%S'),
                    'yoloCount': int(row.get('yoloCount', 0) or 0),
                    'heatmap_path': heatmap_path,
                })
        except Exception:
            pass

    if not camera_frames:
        raise HTTPException(status_code=404, detail=f"No data between {start_dt.strftime('%H:%M:%S')} and {end_dt.strftime('%H:%M:%S')}")

    try:
        path = _build_5min_frame_report(camera_frames, start_dt, end_dt, date_str)
        time_label = start_dt.strftime('%H%M%S')
        return FileResponse(path, media_type="application/pdf",
                            filename=f"IRIS_Crowd_5min_{date_str}_{time_label}.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _parse_camera_spreadsheet(file_bytes: bytes, filename: str) -> list:
    """Parse an uploaded Excel/CSV file and extract camera entries.
    Returns list of {'name': str, 'ip': str, 'rtsp': str}.
    Handles various column layouts by searching for RTSP URLs."""
    cameras = []
    rows_data = []

    if filename.endswith('.csv'):
        import csv, io
        text = file_bytes.decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(text))
        rows_data = [row for row in reader]
    else:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows_data = [[cell for cell in row] for row in ws.iter_rows(min_row=1, values_only=True)]

    if not rows_data or len(rows_data) < 2:
        return cameras

    # Skip header row(s) — find first row with an RTSP URL
    data_start = 1
    for i, row in enumerate(rows_data):
        if any('rtsp://' in str(v).lower() for v in row if v):
            data_start = i
            break

    for row in rows_data[data_start:]:
        vals = [str(v).strip() if v is not None else '' for v in row]
        # Find the RTSP URL column
        rtsp = ''
        rtsp_idx = -1
        for i, v in enumerate(vals):
            if 'rtsp://' in v.lower():
                rtsp = v
                rtsp_idx = i
                break
        if not rtsp:
            continue

        # Extract IP from RTSP URL
        try:
            from urllib.parse import urlparse
            parsed = urlparse(rtsp)
            ip = parsed.hostname or ''
        except Exception:
            ip = ''

        # Find camera name — look for a string column that's not a number and not the IP
        name = ''
        for i, v in enumerate(vals):
            if i == rtsp_idx:
                continue
            if v and not v.replace('.', '').replace('-', '').isdigit() and 'rtsp://' not in v.lower() and len(v) > 3:
                name = v
                break
        if not name:
            name = f"camera_{ip.replace('.', '_')}"

        cameras.append({'name': name, 'ip': ip, 'rtsp': rtsp})

    return cameras


@app.post("/api/crowd-live/cameras/upload")
async def api_crowd_cameras_upload(request: Request):
    """Upload Excel/CSV file, parse cameras, return list for confirmation."""
    import io
    form = await request.form()
    file = form.get('file')
    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded")
    filename = file.filename or 'upload.xlsx'
    file_bytes = await file.read()
    cameras = _parse_camera_spreadsheet(file_bytes, filename)
    if not cameras:
        raise HTTPException(status_code=400, detail="No cameras found in file. Expected columns with camera name and RTSP URL.")
    return {"cameras": cameras, "count": len(cameras)}


@app.post("/api/crowd-live/cameras/apply")
async def api_crowd_cameras_apply(request: Request):
    """Apply parsed camera list to config. Body: {cameras: [...], mode: 'replace'|'append'}"""
    import yaml
    body = await request.json()
    cameras = body.get('cameras', [])
    mode = body.get('mode', 'replace')
    if not cameras:
        raise HTTPException(status_code=400, detail="No cameras provided")

    config_path = Path(__file__).resolve().parent / "config" / "rtsp_links.yml"
    try:
        with open(config_path) as f:
            cfg = yaml.safe_load(f) or {}
    except Exception:
        cfg = {}

    new_entries = [{'name': c['name'], 'url': c['rtsp']} for c in cameras if c.get('rtsp')]

    if mode == 'append':
        existing = cfg.get('crowd_cameras', []) or []
        existing_urls = {e.get('url', '') for e in existing}
        for entry in new_entries:
            if entry['url'] not in existing_urls:
                existing.append(entry)
        final_cameras = existing
    else:
        final_cameras = new_entries

    # Clear cached analysis data from old cameras
    with _crowd_worker_lock:
        _crowd_worker_analysis.clear()
        _crowd_worker_history.clear()

    # Write YAML preserving structure (rtsp_links first, then crowd_cameras)
    lines = []
    lines.append("rtsp_links:")
    for url in cfg.get('rtsp_links', []):
        lines.append(f"- {url}")
    lines.append("active_sources: []")
    lines.append("overlays: {}")
    lines.append("crowd_cameras:")
    for cam in final_cameras:
        lines.append(f"- name: {cam['name']}")
        lines.append(f"  url: {cam['url']}")
    lines.append("crowd_camera_groups: []")
    lines.append("")
    with open(config_path, 'w') as f:
        f.write('\n'.join(lines))

    # Restart crowd worker
    import signal, subprocess
    try:
        result = subprocess.run(['pgrep', '-f', 'crowd_worker.py'], capture_output=True, text=True)
        for pid in result.stdout.strip().split('\n'):
            if pid.strip():
                try:
                    os.kill(int(pid.strip()), signal.SIGKILL)
                except ProcessLookupError:
                    pass
        import threading
        def _restart_worker():
            import time as _t
            _t.sleep(2)
            venv_python = str(Path(__file__).resolve().parent / ".venv" / "bin" / "python")
            worker_script = str(Path(__file__).resolve().parent / "crowd_worker.py")
            subprocess.Popen([venv_python, "-u", worker_script],
                             stdout=open('/tmp/crowd_worker.log', 'w'),
                             stderr=subprocess.STDOUT,
                             cwd=str(Path(__file__).resolve().parent),
                             start_new_session=True)
            print(f"[CW] Worker restarted with {len(final_cameras)} cameras")
        threading.Thread(target=_restart_worker, daemon=True).start()
    except Exception as e:
        print(f"[CW] Worker restart failed: {e}")

    return {
        "status": "ok",
        "mode": mode,
        "cameras_count": len(cfg.get('crowd_cameras', [])),
        "message": f"{'Replaced' if mode == 'replace' else 'Appended'} {len(new_entries)} cameras. Worker restarting."
    }


@app.get("/api/crowd-live/stream/{camera_id}")
def api_crowd_live_stream(camera_id: str):
    """MJPEG stream of live heatmap-overlaid inference frames for a camera."""
    def generate():
        while True:
            frame_data = rt_crowd.get_live_cam_frame(camera_id)
            if frame_data:
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_data + b'\r\n')
            time.sleep(0.5)
    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@app.get("/api/crowd-live/analysis")
def api_crowd_live_all_analysis():
    """Per-camera analysis data for all cameras (JSON, no frames)."""
    return rt_crowd.get_all_cam_analysis()

@app.get("/api/crowd-live/analysis/{camera_id}")
def api_crowd_live_camera_analysis(camera_id: str):
    """Per-camera analysis data for a single camera (JSON, no frame)."""
    data = rt_crowd.get_live_cam_analysis_data(camera_id)
    if not data:
        raise HTTPException(status_code=404, detail="No analysis data for this camera")
    return data


# ── Crowd-worker endpoints (external crowd-worker.py integration) ──

@app.get("/api/crowd-worker/config")
def api_crowd_worker_config():
    """Return camera list from rtsp_links.yml formatted for crowd-worker.py."""
    import yaml
    from urllib.parse import urlparse
    config_path = Path(__file__).resolve().parent / "config" / "rtsp_links.yml"
    try:
        with open(config_path) as f:
            cfg = yaml.safe_load(f) or {}
    except Exception:
        return {"tasks": []}
    tasks = []
    # Only serve crowd_cameras for the worker
    for cam in cfg.get("crowd_cameras", []):
        name = cam.get("name", "")
        url = cam.get("url", "")
        if url:
            tasks.append({"id": name or f"crowd_{urlparse(url).hostname}", "rtspUrl": url, "name": name})
    return {"tasks": tasks}

@app.post("/api/crowd/analysis")
async def api_crowd_analysis_post(request: Request):
    """Receive analysis data from crowd-worker process."""
    global _crowd_worker_session_start
    data = await request.json()
    device_id = data.get("deviceId", "")
    if not device_id:
        raise HTTPException(400, "Missing deviceId")
    with _crowd_worker_lock:
        _crowd_worker_analysis[device_id] = data
        # Track session start
        if not _crowd_worker_session_start:
            _crowd_worker_session_start = datetime.now().isoformat()
        # Accumulate history for reports
        _crowd_worker_history.setdefault(device_id, []).append(data)
        # Keep max 200 snapshots per camera per segment
        if len(_crowd_worker_history[device_id]) > 200:
            _crowd_worker_history[device_id] = _crowd_worker_history[device_id][-200:]
        # Check if segment should close
        _cw_maybe_close_segment()
    # Log to CSV (outside lock)
    _cw_append_csv(data)
    return {"ok": True}

@app.get("/api/crowd/analysis/latest")
def api_crowd_analysis_latest():
    """Return latest crowd-worker analysis data for all cameras.

    Filters the in-memory cache against the current `crowd_cameras` set in
    rtsp_links.yml — so a worker restart that races a config change can't
    leak ghost tiles for cameras that are no longer configured.
    """
    import yaml
    config_path = Path(__file__).resolve().parent / "config" / "rtsp_links.yml"
    try:
        with open(config_path) as f:
            cfg = yaml.safe_load(f) or {}
        active_ids = {
            (cam.get("name") or "").strip()
            for cam in (cfg.get("crowd_cameras") or [])
            if isinstance(cam, dict) and (cam.get("name") or "").strip()
        }
    except Exception:
        active_ids = None  # if we can't read config, don't filter

    with _crowd_worker_lock:
        snapshot = dict(_crowd_worker_analysis)

    if active_ids is None:
        return snapshot
    return {cid: data for cid, data in snapshot.items() if cid in active_ids}

@app.get("/api/heatmaps/{filename:path}")
def api_serve_heatmap(filename: str):
    """Serve heatmap images saved by crowd-worker."""
    filepath = os.path.join(_HEATMAP_DIR, filename)
    if not os.path.isfile(filepath):
        raise HTTPException(404, "Heatmap not found")
    return FileResponse(filepath, media_type="image/jpeg")


# ── Upload endpoints ──

upload_sources_lock = threading.Lock()
upload_sources: Dict[str, dict] = {}
upload_name_counter = 0
upload_chunk_sessions_lock = threading.Lock()
upload_chunk_sessions: Dict[str, dict] = {}
DEFAULT_UPLOAD_CHUNK_SIZE = int(os.environ.get("IRIS_UPLOAD_CHUNK_SIZE", str(25 * 1024 * 1024)))
MAX_UPLOAD_CHUNK_SIZE = int(os.environ.get("IRIS_MAX_UPLOAD_CHUNK_SIZE", str(100 * 1024 * 1024)))
UPLOAD_SESSION_TTL_SECONDS = int(os.environ.get("IRIS_UPLOAD_SESSION_TTL_SECONDS", "7200"))


def _queue_forensics_job_from_upload(file_path: str, display_name: str) -> Optional[str]:
    """Create a Gemini forensics job from an existing upload file without touching the source file."""
    try:
        src = Path(file_path)
        if not src.exists():
            print(f"[FORENSICS] Cannot queue job, upload file missing: {file_path}")
            return None

        safe_name = Path(display_name or src.name).name
        # Default to scoped cancellation (same video name) to avoid unrelated upload jobs
        # being cancelled whenever a new upload/restart queues Gemini forensics.
        cancel_all = os.environ.get("IRIS_FORENSICS_CANCEL_ACTIVE_ON_NEW_UPLOAD", "0").strip().lower() in {"1", "true", "yes", "on"}
        if cancel_all:
            cancelled = forensics_events.cancel_active_jobs(reason=f"Superseded by new upload: {safe_name}")
        else:
            cancelled = forensics_events.cancel_active_jobs(video_name=safe_name, reason=f"Superseded by newer upload run: {safe_name}")
        if cancelled:
            print(f"[FORENSICS] Cancelled {len(cancelled)} active Gemini job(s) before queueing new upload")

        fast_link_default = "1" if (IRIS_LOCAL or int(os.environ.get("IRIS_BACKEND_PORT", "9010")) == 19010) else "0"
        fast_link = os.environ.get("IRIS_FORENSICS_FAST_LINK_UPLOAD", fast_link_default).strip().lower() in {"1", "true", "yes", "on"}
        if fast_link:
            # Dev-only fast path: avoid extra file copy and analyze upload source directly.
            job_id = forensics_events.create_job_from_path(safe_name, str(src))
        else:
            suffix = src.suffix or ".mp4"
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=suffix,
                dir=str(FORENSICS_UPLOAD_DIR),
                prefix="linked_",
            ) as tmp:
                tmp_path = Path(tmp.name)
            shutil.copy2(src, tmp_path)
            job_id = forensics_events.create_job_from_path(safe_name, str(tmp_path))
        print(f"[FORENSICS] Queued Gemini job {job_id} from upload source={src.name}")
        return job_id
    except Exception as e:
        print(f"[FORENSICS] Failed to queue job from upload: {e}")
        return None
def _next_upload_name() -> str:
    global upload_name_counter
    with upload_sources_lock:
        upload_name_counter += 1
        return f"upload{upload_name_counter}"


def _validate_upload_format(filename: str, content_type: str = ""):
    ext = Path(filename or "").suffix.lower()
    allowed_exts = {".mp4", ".mkv", ".avi", ".mov"}
    if ext in allowed_exts:
        return
    ctype = (content_type or "").lower()
    allowed_mimes = {"video/mp4", "video/x-matroska", "video/quicktime", "video/x-msvideo"}
    if ctype not in allowed_mimes and not ctype.startswith("video/"):
        raise HTTPException(400, "Unsupported file format. Use MP4, MKV, AVI, or MOV.")


def _purge_stale_upload_chunk_sessions():
    now = time.time()
    stale = []
    with upload_chunk_sessions_lock:
        for upload_id, info in upload_chunk_sessions.items():
            if now - float(info.get("updated_at", now)) > UPLOAD_SESSION_TTL_SECONDS:
                stale.append((upload_id, info.get("file_path")))
        for upload_id, _ in stale:
            upload_chunk_sessions.pop(upload_id, None)
    for _, file_path in stale:
        if not file_path:
            continue
        try:
            Path(file_path).unlink(missing_ok=True)
        except Exception:
            pass


def _start_uploaded_recording(file_path: Path, original_filename: str, mode: Optional[str] = None):
    _validate_upload_format(original_filename)
    _ensure_inference_can_start()

    global active_mode
    requested_mode = normalize_mode(mode)
    upload_mode = requested_mode or normalize_mode(active_mode) or "vehicle"
    if not requested_mode and not normalize_mode(active_mode):
        print("[UPLOAD] No route mode context found; defaulting upload mode to vehicle (VCC).")
    if requested_mode and active_mode and requested_mode != active_mode:
        print(f"[MODE] Running mixed modes in parallel: global={active_mode}, upload={requested_mode}")
    elif requested_mode and not active_mode:
        active_mode = requested_mode

    original_name = Path(original_filename).stem
    job_id = uuid.uuid4().hex
    name = _next_upload_name()
    ensure_overlay(name)

    # If a previous upload with the same original name exists, stop and remove it
    with upload_sources_lock:
        existing = [n for n, info in upload_sources.items() if info.get("original_name") == original_name]
    for n in existing:
        _stop_upload_by_name(n, delete_file=False)

    duration = get_video_duration(str(file_path))
    source_fps = get_video_fps(str(file_path))
    if source_fps <= 0:
        source_fps = float(PROCESSED_FPS)

    base_overlay = route_mode_overlay(upload_mode)
    mode_confidence = route_mode_confidence(upload_mode)
    overlay_config = {**base_overlay, "confidence": mode_confidence, "active_mode": upload_mode}

    with overlay_lock:
        overlays[name] = dict(overlay_config)

    is_crowd_mode = upload_mode == "crowd"
    print(f"[UPLOAD] {name} overlays from mode={upload_mode}, is_crowd={is_crowd_mode}: {overlay_config}")

    # Calculate total active streams for dynamic GPU allocation
    with upload_sources_lock:
        upload_count = len(upload_sources)
    active_streams = len(running_sources) + upload_count + 1

    if upload_count >= MAX_UPLOAD_STREAMS:
        raise HTTPException(429, f"Maximum upload streams reached ({MAX_UPLOAD_STREAMS}).")

    # Use configurable realtime mode for uploads; default is fast mode for lower latency.
    process, stop = start_upload_callback(
        str(file_path),
        name,
        overlay_config,
        is_crowd_mode,
        active_streams,
        realtime=UPLOAD_REALTIME,
    )
    if process is None:
        raise HTTPException(500, "Failed to start upload inference")
    if UPLOAD_RAW_PASSTHROUGH:
        _start_upload_raw_passthrough_async(name, str(file_path))
    else:
        _start_raw_ffmpeg_publisher_async(name)

    with upload_sources_lock:
        upload_sources[name] = {
            "process": process,
            "stop": stop,
            "file_path": str(file_path),
            "job_id": job_id,
            "original_name": original_name,
            "duration": duration,
            "source_fps": source_fps,
            "started_processing": False,
            "started_at": None,
            "mode": upload_mode,
            "realtime": UPLOAD_REALTIME,
            "forensics_job_id": None,
        }

    # Queue Gemini forensics for all upload modes so the comprehensive report is available.
    forensics_job_id = _queue_forensics_job_from_upload(str(file_path), original_filename)
    if forensics_job_id:
        with upload_sources_lock:
            if name in upload_sources:
                upload_sources[name]["forensics_job_id"] = forensics_job_id

    with jobs_lock:
        jobs[job_id] = {"id": job_id, "name": name, "status": "processing"}

    _ffmpeg_monitor_sources.add(name)
    ffmpeg_next_start[name] = time.time() + 0.15
    _start_ffmpeg_publisher_async(name)
    _ensure_ffmpeg_monitor()

    return {
        "job_id": job_id,
        "name": name,
        "status": "started",
        "forensics_job_id": forensics_job_id,
    }

def _stop_all_uploads(delete_files: bool = False):
    """Stop all upload processes. Optionally delete files and entries."""
    with upload_sources_lock:
        names = list(upload_sources.keys())

    for name in names:
        _ffmpeg_monitor_sources.discard(name)
        _stop_ffmpeg_publisher(name)
        _stop_raw_ffmpeg_publisher(name)

    with upload_sources_lock:
        for name, info in list(upload_sources.items()):
            try:
                f_job = info.get("forensics_job_id")
                if f_job:
                    if forensics_events.cancel_job(f_job, reason=f"Stop-all: {name}"):
                        print(f"[FORENSICS] Cancelled Gemini job {f_job} for {name} on stop-all")
            except Exception:
                pass
            info["stop"].set()
            proc = info.get("process")
            if proc:
                try:
                    proc.terminate()
                    proc.join(timeout=2)
                except:
                    pass
            info["process"] = None
            info["stopped"] = True
            if delete_files:
                try:
                    Path(info["file_path"]).unlink()
                except:
                    pass
                upload_sources.pop(name, None)

    # Clear per-upload buffers/metrics/overlays (runtime state only)
    with metrics_lock:
        for name in names:
            metrics.pop(name, None)
    with overlay_lock:
        for name in names:
            overlays.pop(name, None)
    with frame_lock:
        for name in names:
            frame_buffer.pop(name, None)
            frame_bgr_buffer.pop(name, None)
            frame_sequences.pop(name, 0)
    with raw_frame_lock:
        for name in names:
            raw_frame_buffer.pop(name, None)
            raw_frame_bgr_buffer.pop(name, None)
            raw_frame_sequences.pop(name, None)

def _stop_upload_by_name(name: str, delete_file: bool = False):
    _ffmpeg_monitor_sources.discard(name)
    _stop_ffmpeg_publisher(name)
    _stop_raw_ffmpeg_publisher(name)
    with upload_sources_lock:
        info = upload_sources.get(name)
        if not info:
            return
        # Cancel by job ID directly (most reliable — no name-matching ambiguity)
        try:
            f_job = info.get("forensics_job_id")
            if f_job:
                if forensics_events.cancel_job(f_job, reason=f"Upload stopped: {name}"):
                    print(f"[FORENSICS] Cancelled Gemini job {f_job} for {name}")
                frames_dir = forensics_events.get_frames_dir(f_job)
                if frames_dir:
                    (frames_dir / "vcc_complete.txt").write_text("done")
        except Exception:
            pass
        info["stop"].set()
        proc = info.get("process")
        if proc:
            try:
                proc.terminate()
                proc.join(timeout=2)
            except Exception:
                pass
        if delete_file:
            try:
                Path(info["file_path"]).unlink()
            except Exception:
                pass
        upload_sources.pop(name, None)

    with metrics_lock:
        metrics.pop(name, None)
    with overlay_lock:
        overlays.pop(name, None)
    with frame_lock:
        frame_buffer.pop(name, None)
        frame_bgr_buffer.pop(name, None)
        frame_sequences.pop(name, 0)
    with raw_frame_lock:
        raw_frame_buffer.pop(name, None)
        raw_frame_bgr_buffer.pop(name, None)
        raw_frame_sequences.pop(name, None)

@app.post("/api/upload")
async def upload_video(file: UploadFile = File(...), mode: Optional[str] = Form(None)):
    filename = Path(file.filename or "upload.mp4").name

    # Ensure directory exists in case it was deleted at runtime
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    target = UPLOAD_DIR / f"{uuid.uuid4().hex}_{filename}"
    with open(target, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return _start_uploaded_recording(target, filename, mode)


@app.post("/api/upload/init")
def upload_init(req: UploadInitRequest):
    _purge_stale_upload_chunk_sessions()
    filename = Path(req.filename or "upload.mp4").name
    _validate_upload_format(filename)
    if req.size <= 0:
        raise HTTPException(400, "Invalid upload size")

    requested_chunk = int(req.chunk_size or DEFAULT_UPLOAD_CHUNK_SIZE)
    chunk_size = min(max(1 * 1024 * 1024, requested_chunk), MAX_UPLOAD_CHUNK_SIZE)
    total_chunks = (req.size + chunk_size - 1) // chunk_size

    upload_id = uuid.uuid4().hex
    target = UPLOAD_DIR / f"{upload_id}_{filename}"
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    target.touch(exist_ok=False)

    with upload_chunk_sessions_lock:
        upload_chunk_sessions[upload_id] = {
            "file_path": str(target),
            "filename": filename,
            "size": int(req.size),
            "chunk_size": int(chunk_size),
            "total_chunks": int(total_chunks),
            "mode": normalize_mode(req.mode),
            "received": {},  # index -> bytes written
            "updated_at": time.time(),
        }

    return {"upload_id": upload_id, "chunk_size": int(chunk_size), "total_chunks": int(total_chunks)}


@app.post("/api/upload/chunk")
async def upload_chunk(request: Request, upload_id: str, index: int):
    if index < 0:
        raise HTTPException(400, "Invalid chunk index")
    with upload_chunk_sessions_lock:
        info = upload_chunk_sessions.get(upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        chunk_size = int(info["chunk_size"])
        total_chunks = int(info["total_chunks"])
        file_path = str(info["file_path"])
    if index >= total_chunks:
        raise HTTPException(400, "Chunk index out of range")

    body = await request.body()
    if not body:
        raise HTTPException(400, "Empty chunk")

    offset = index * chunk_size
    with open(file_path, "r+b") as f:
        f.seek(offset)
        f.write(body)

    with upload_chunk_sessions_lock:
        info = upload_chunk_sessions.get(upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        received = info.setdefault("received", {})
        received[index] = len(body)
        info["updated_at"] = time.time()
        done = len(received) >= int(info["total_chunks"])

    return {"ok": True, "done": done}


@app.post("/api/upload/complete")
def upload_complete(req: UploadCompleteRequest):
    with upload_chunk_sessions_lock:
        info = upload_chunk_sessions.get(req.upload_id)
        if not info:
            raise HTTPException(404, "Upload session not found")
        file_path = Path(str(info["file_path"]))
        filename = str(info["filename"])
        mode = info.get("mode")
        total_chunks = int(info["total_chunks"])
        expected_size = int(info["size"])
        received = dict(info.get("received", {}))
        upload_chunk_sessions.pop(req.upload_id, None)

    if len(received) != total_chunks:
        raise HTTPException(400, "Upload incomplete")
    if sum(int(v) for v in received.values()) != expected_size:
        raise HTTPException(400, "Upload size mismatch")
    if not file_path.exists():
        raise HTTPException(400, "Uploaded file missing")

    return _start_uploaded_recording(file_path, filename, mode)

@app.post("/api/uploads/{name}/restart")
def restart_upload(name: str, mode: Optional[str] = None):
    """Stop and restart an existing upload."""
    _ensure_inference_can_start()
    print(f"[UPLOAD] Restarting {name} (requested_mode={mode})")
    
    # 1. Stop if running
    with upload_sources_lock:
        if name in upload_sources:
            info = upload_sources[name]
            info["stop"].set()
            if "process" in info:
                try:
                    info["process"].join(timeout=2)
                except:
                    pass
            # Don't pop yet, we need the file path
            file_path = info["file_path"]
            original_name = info.get("original_name")
            job_id = info.get("job_id")
            info["stopped"] = False
        else:
            raise HTTPException(404, "Upload not active or found")

    # 2. Start again
    # Route-mode override prevents cross-route processed-feed poisoning.
    upload_mode = None
    with upload_sources_lock:
        upload_mode = upload_sources.get(name, {}).get("mode")
    restart_mode = normalize_mode(mode) or normalize_mode(upload_mode) or normalize_mode(active_mode)
    if not restart_mode:
        raise HTTPException(400, "No route selected. Open a route in UI before restarting inference.")

    base_overlay = route_mode_overlay(restart_mode)
    mode_confidence = route_mode_confidence(restart_mode)
    overlay_config = {**base_overlay, "confidence": mode_confidence, "active_mode": restart_mode}
    
    is_crowd_mode = restart_mode == "crowd"
    
    # Update overlays dict
    with overlay_lock:
        overlays[name] = dict(overlay_config)

    # Recalculate streams count
    with source_lock:
        rtsp_count = len(running_sources)
    with upload_sources_lock:
        # It's still in the dict, so count is same
        upload_count = len(upload_sources)
    
    active_streams = rtsp_count + upload_count

    # Cleanup old process info in ffmpeg monitor just in case
    _ffmpeg_monitor_sources.discard(name)
    _stop_ffmpeg_publisher(name)
    _stop_raw_ffmpeg_publisher(name)

    # Start new process
    realtime_mode = bool(upload_sources.get(name, {}).get("realtime", UPLOAD_REALTIME))
    process, stop = start_upload_callback(
        file_path,
        name,
        overlay_config,
        is_crowd_mode,
        active_streams,
        realtime=realtime_mode,
    )
    if process is None:
        raise HTTPException(500, "Failed to restart upload inference")
    if UPLOAD_RAW_PASSTHROUGH:
        _start_upload_raw_passthrough_async(name, file_path)
    else:
        _start_raw_ffmpeg_publisher_async(name)

    with upload_sources_lock:
        prev = upload_sources.get(name, {})
        upload_sources[name] = {
            "process": process,
            "stop": stop,
            "file_path": file_path,
            "job_id": job_id,
            "original_name": original_name,
            "source_fps": float(prev.get("source_fps", PROCESSED_FPS)),
            "started_at": None,
            "mode": restart_mode,
            "realtime": realtime_mode,
            "forensics_job_id": prev.get("forensics_job_id"),
        }
    
    # Reset job status
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["status"] = "processing"

    # Restart monitor
    _ffmpeg_monitor_sources.add(name)
    ffmpeg_next_start[name] = time.time() + 0.15
    _start_ffmpeg_publisher_async(name)
    _ensure_ffmpeg_monitor()

    # Reuse an existing non-terminal Gemini job for this upload; restarting video inference
    # should not continuously supersede/cancel forensics work.
    existing_forensics_job_id = None
    with upload_sources_lock:
        existing_forensics_job_id = upload_sources.get(name, {}).get("forensics_job_id")
    forensics_job_id = None
    if existing_forensics_job_id:
        existing_status = forensics_events.get_status(str(existing_forensics_job_id)) or {}
        existing_state = str(existing_status.get("status", "")).strip().lower()
        if existing_state and existing_state not in {"error", "cancelled"}:
            forensics_job_id = str(existing_forensics_job_id)
            print(f"[FORENSICS] Reusing existing Gemini job {forensics_job_id} for upload {name} (status={existing_state})")

    # Queue Gemini forensics only when there is no reusable active/completed job.
    if not forensics_job_id:
        display_name = (original_name or name) + Path(file_path).suffix
        forensics_job_id = _queue_forensics_job_from_upload(str(file_path), display_name)
    if forensics_job_id:
        with upload_sources_lock:
            if name in upload_sources:
                upload_sources[name]["forensics_job_id"] = forensics_job_id

    return {"status": "restarted", "name": name, "mode": restart_mode, "forensics_job_id": forensics_job_id}

@app.get("/api/uploads")
def list_uploads():
    with upload_sources_lock:
        return {
            "uploads": [
                {
                    "name": name,
                    "original_name": info.get("original_name", name),
                    "job_id": info.get("job_id"),
                    "forensics_job_id": info.get("forensics_job_id"),
                    "mode": info.get("mode"),
                    "source_fps": float(info.get("source_fps", 0.0) or 0.0),
                    "stream_url": f"/api/stream/{name}",
                    "stopped": info.get("stopped", False),
                }
                for name, info in upload_sources.items()
            ]
        }

@app.post("/api/uploads/{name}/stop")
def stop_upload(name: str):
    _ffmpeg_monitor_sources.discard(name)
    _stop_ffmpeg_publisher(name)
    _stop_raw_ffmpeg_publisher(name)
    with upload_sources_lock:
        info = upload_sources.get(name)
        if info:
            # Cancel by job ID directly — most reliable
            try:
                f_job = info.get("forensics_job_id")
                if f_job:
                    if forensics_events.cancel_job(f_job, reason=f"Upload stopped by user: {name}"):
                        print(f"[FORENSICS] Cancelled Gemini job {f_job} for {name}")
            except Exception:
                pass
            info["stop"].set()
            proc = info.get("process")
            if proc:
                try:
                    proc.terminate()
                    proc.join(timeout=2)
                except Exception:
                    pass
            info["process"] = None
            info["stopped"] = True
            return {"status": "stopped", "name": name}
    raise HTTPException(404, "Upload not found")

@app.delete("/api/uploads/{name}")
def delete_upload(name: str):
    _ffmpeg_monitor_sources.discard(name)
    _stop_ffmpeg_publisher(name)
    _stop_raw_ffmpeg_publisher(name)
    with upload_sources_lock:
        info = upload_sources.pop(name, None)
        if info:
            try:
                f_job = info.get("forensics_job_id")
                if f_job:
                    if forensics_events.cancel_job(f_job, reason=f"Upload deleted by user: {name}"):
                        print(f"[FORENSICS] Cancelled Gemini job {f_job} for deleted upload {name}")
            except Exception:
                pass
            info["stop"].set()
            proc = info.get("process")
            if proc:
                try:
                    proc.terminate()
                    proc.join(timeout=2)
                except:
                    pass
            try:
                Path(info["file_path"]).unlink()
            except:
                pass
            return {"status": "deleted", "name": name}
    raise HTTPException(404, "Upload not found")


# ── SAM3 endpoints ──

@app.post("/api/sam/start")
def sam_start(req: SamStartRequest):
    _ensure_inference_can_start()
    if not load_sam_model():
        raise HTTPException(503, "SAM3 model failed to load")

    existing = sam_threads.pop(req.source, None)
    if existing:
        existing["stop_event"].set()
        # Non-blocking or short-timeout join to avoid hanging API
        # existing["thread"].join(timeout=1) 

    with sam_results_lock:
        sam_results[req.source] = {"session_history": [], "vlm_analysis": None}

    settings_ref = {
        "confidence": req.confidence,
        "show_boxes": req.show_boxes,
        "show_masks": req.show_masks,
    }

    stop_event = threading.Event()
    t = threading.Thread(
        target=sam_worker,
        args=(req.source, req.prompt, req.confidence, stop_event,
              req.show_boxes, req.show_masks, settings_ref),
        kwargs={"raw_frame_lock": raw_frame_lock, "raw_frame_buffer": raw_frame_buffer},
        daemon=True,
    )
    t.start()
    sam_threads[req.source] = {
        "thread": t,
        "stop_event": stop_event,
        "prompt": req.prompt,
        "confidence": req.confidence,
        "show_boxes": req.show_boxes,
        "show_masks": req.show_masks,
        "settings_ref": settings_ref,
    }
    return {"status": "started", "source": req.source, "prompt": req.prompt}

@app.post("/api/sam/update")
def sam_update(req: SamUpdateRequest):
    info = sam_threads.get(req.source)
    if not info:
        raise HTTPException(404, "SAM not running for this source")

    settings_ref = info.get("settings_ref")
    if not settings_ref:
        raise HTTPException(500, "Settings ref not available")

    if req.confidence is not None:
        settings_ref["confidence"] = req.confidence
        info["confidence"] = req.confidence
    if req.show_boxes is not None:
        settings_ref["show_boxes"] = req.show_boxes
        info["show_boxes"] = req.show_boxes
    if req.show_masks is not None:
        settings_ref["show_masks"] = req.show_masks
        info["show_masks"] = req.show_masks

    return {
        "status": "updated",
        "source": req.source,
        "confidence": settings_ref["confidence"],
        "show_boxes": settings_ref["show_boxes"],
        "show_masks": settings_ref["show_masks"],
    }

@app.get("/api/sam/result/{source}")
def sam_result(source: str):
    with sam_results_lock:
        result = sam_results.get(source)
    if result is None:
        return {"status": "no_result", "source": source}
    return result

@app.post("/api/sam/stop")
def sam_stop(req: SamStopRequest):
    info = sam_threads.pop(req.source, None)
    if info:
        info["stop_event"].set()
        return {"status": "stopped", "source": req.source}
    return {"status": "not_running", "source": req.source}

@app.get("/api/sam/status")
def sam_status():
    from sam import sam_model_loaded as _loaded
    active = list(sam_threads.keys())
    return {
        "model_loaded": _loaded,
        "active_sources": active,
        "details": {
            src: {
                "prompt": info["prompt"],
                "confidence": info["confidence"],
                "show_boxes": info.get("show_boxes", True),
                "show_masks": info.get("show_masks", True),
            }
            for src, info in sam_threads.items()
        },
    }


# ── Server entry point ──

def run_control_server(start_source_fn, start_upload_fn, initial_overlays):
    global start_source_callback, start_upload_callback, overlays

    start_source_callback = start_source_fn
    start_upload_callback = start_upload_fn
    forensics_events.load_jobs_from_disk()

    if initial_overlays is not None:
        overlays = initial_overlays
        print(f"[OVERLAY] Using shared Manager dict (type: {type(initial_overlays).__name__})")
    else:
        overlays = {}
        print("[OVERLAY] Using local dict (no shared dict provided)")

    cfg = load_rtsp_config()
    rtsp_links = cfg.get("rtsp_links", [])

    # Clear persisted runtime state so each run starts clean.
    if cfg.get("active_sources") or cfg.get("overlays"):
        cfg["active_sources"] = []
        cfg["overlays"] = {}
        save_rtsp_config(cfg)

    saved_overlays = {}
    for name, state in saved_overlays.items():
        with overlay_lock:
            heatmap_val = state.get("heatmap", True)
            overlays[name] = {
                "heatmap": heatmap_val,
                "heatmap_full": state.get("heatmap_full", heatmap_val),
                "heatmap_trails": state.get("heatmap_trails", heatmap_val),
                "trails": state.get("trails", True),
                "bboxes": state.get("bboxes", True),
                "confidence": state.get("confidence", 0.15),
            }
            print(f"[OVERLAY] Loaded saved state for {name}: {overlays[name]}")

    # Do not pre-initialize overlays for all configured RTSP links on boot.
    # Overlays are created lazily when a source/upload is used.
    print(
        f"[STARTUP] Control server initialized | rtsp_links_configured={len(rtsp_links)} "
        f"| lazy_overlay_init=1"
    )

    class _ShutdownNoiseFilter(logging.Filter):
        def filter(self, record: logging.LogRecord) -> bool:
            if record.exc_info:
                exc_type = record.exc_info[0]
                if exc_type in (asyncio.CancelledError, KeyboardInterrupt):
                    return False
            return True

    logging.getLogger("uvicorn.error").addFilter(_ShutdownNoiseFilter())

    import uvicorn
    import copy
    import threading
    import time

    # --- Access Log Filtering ---
    class AccessLogFilter(logging.Filter):
        def filter(self, record: logging.LogRecord) -> bool:
            if record.args and len(record.args) >= 5:
                status_code = record.args[4]
                path = record.args[2]
                
                if status_code == 200:
                    if "/api/health" in path:
                        return False
                    if path.startswith("/api/sources"):
                        return False
                    if "/api/sam/result/" in path or "/api/sam/status" in path:
                         return False
                    if "/api/uploads" in path:
                         return False
            return True


    # --- Periodic Status Logger ---
    def _status_logger_worker():
        while True:
            time.sleep(15)
            try:
                # Count resources
                with source_lock:
                    n_rtsp = len(running_sources)
                with upload_sources_lock:
                    n_upload = sum(1 for u in upload_sources.values() if u.get("started_processing"))
                n_sam = len(sam_threads)
                
                print(f"[SYSTEM] Status: OK | RTSP Streams: {n_rtsp} | Uploads: {n_upload} | SAM Workers: {n_sam}")
            except Exception:
                pass

    threading.Thread(target=_status_logger_worker, daemon=True).start()

    def _memory_guard_worker():
        if not MEMORY_GUARD_ENABLED:
            print("[MEMORY_GUARD] Disabled")
            return
        print(
            f"[MEMORY_GUARD] Enabled | threshold={MEMORY_GUARD_AVAIL_MB}MB | "
            f"cooldown={MEMORY_GUARD_COOLDOWN_SECONDS}s | poll={MEMORY_GUARD_POLL_SECONDS}s"
        )
        while True:
            time.sleep(max(0.5, MEMORY_GUARD_POLL_SECONDS))
            try:
                vm = psutil.virtual_memory()
                avail_mb = int(vm.available / (1024 * 1024))
                if avail_mb > MEMORY_GUARD_AVAIL_MB:
                    continue

                # Avoid repeated stop storms while already in cooldown.
                if _inference_pause_remaining_seconds() > 0:
                    continue

                reason = f"low memory ({avail_mb}MB <= {MEMORY_GUARD_AVAIL_MB}MB)"
                resume_at = _set_inference_pause(MEMORY_GUARD_COOLDOWN_SECONDS, reason)
                stopped = _stop_all_sources()
                wait_s = max(1, int(math.ceil(resume_at - time.time())))
                print(
                    f"[MEMORY_GUARD] Triggered: {reason}. "
                    f"Stopped inference sources={stopped}. New starts blocked for {wait_s}s."
                )
            except Exception as e:
                print(f"[MEMORY_GUARD] Error: {e}")

    threading.Thread(target=_memory_guard_worker, daemon=True).start()

    # Define standard Uvicorn log config
    log_config = uvicorn.config.LOGGING_CONFIG.copy()
    
    # Add our filter to the configuration
    log_config["filters"] = log_config.get("filters", {})
    log_config["filters"]["access_filter"] = {
        "()": lambda: AccessLogFilter()
    }
    
    # Apply filter to access logger
    log_config["loggers"]["uvicorn.access"]["filters"] = ["access_filter"]

    # Run with explicit log_config; port is configurable for dev/prod isolation.
    backend_port = int(os.environ.get("IRIS_BACKEND_PORT", "9010"))
    uvicorn.run(app, host="0.0.0.0", port=backend_port, log_config=log_config)


# ── MagicBox Crowd (Edge Device Head Counts) ─────────────────────────────
_mb_crowd: dict = {}          # {device_id: {last_report, first_seen, last_seen, total_reports}}
_mb_crowd_lock = threading.Lock()
_mb_crowd_frames = deque(maxlen=500)  # recent frames for /frames API
_mb_fleet_cache: dict = {"data": None, "ts": 0}
_MB_FRAMES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "crowd_frames")
import base64 as _b64


import collections as _collections
_mb_jpeg_queue = _collections.deque()  # FIFO queue of (timestamp, filepath)
_mb_jpeg_queue_loaded = [False]

def _save_frame_to_disk(device_id, camera_name, head_count, timestamp_str, frame_b64):
    """Save ALL frames as JPEG + metadata. FIFO: after 3hrs of data, each new frame
    deletes the oldest one. Metadata JSONL + CSV kept permanently for UI stats."""
    try:
        from datetime import timezone as _tz3, timedelta as _td3
        _IST3 = _tz3(_td3(hours=5, minutes=30))
        _UTC = _tz3(_td3(0))

        ts_clean = timestamp_str.replace("Z", "+00:00") if timestamp_str.endswith("Z") else timestamp_str
        try:
            dt = datetime.fromisoformat(ts_clean)
        except Exception:
            dt = datetime.utcnow()
        # Always ensure timezone-aware (fix naive UTC fallback)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_UTC)
        dt_ist = dt.astimezone(_IST3)
        day_str = dt_ist.strftime("%Y-%m-%d")
        day_dir = os.path.join(_MB_FRAMES_DIR, day_str)
        os.makedirs(day_dir, exist_ok=True)
        safe_dev = device_id.replace(".", "-")
        safe_cam = re.sub(r"[^a-zA-Z0-9_-]", "_", camera_name)[:30]
        ts_file = dt_ist.strftime("%H-%M-%S")
        # Add microsecond hash to prevent filename collisions between cameras
        import hashlib as _hl
        cam_hash = _hl.md5(camera_name.encode()).hexdigest()[:4]
        fname = f"{ts_file}_{safe_dev}_{cam_hash}_{head_count}.jpg"

        # Save JPEG
        fpath = os.path.join(day_dir, fname)
        jpg_bytes = _b64.b64decode(frame_b64)
        with open(fpath, "wb") as f:
            f.write(jpg_bytes)

        # Bootstrap queue on first call (load existing JPEGs from disk)
        if not _mb_jpeg_queue_loaded[0]:
            _mb_jpeg_queue_loaded[0] = True
            import glob as _glob
            cutoff_boot = (dt_ist - _td3(hours=3)).strftime("%Y-%m-%dT%H:%M:%S")
            for ddir in sorted(os.listdir(_MB_FRAMES_DIR)):
                dpath = os.path.join(_MB_FRAMES_DIR, ddir)
                if os.path.isdir(dpath) and len(ddir) == 10 and ddir[4] == "-":
                    for jf in sorted(_glob.glob(os.path.join(dpath, "*.jpg"))):
                        bname = os.path.basename(jf)
                        ts_key = ddir + "T" + bname[:8].replace("-", ":")
                        if ts_key < cutoff_boot:
                            # Orphan older than 3hrs — clean it up
                            try:
                                os.remove(jf)
                            except Exception:
                                pass
                        else:
                            _mb_jpeg_queue.append((ts_key, jf))
            print(f"[magicbox-crowd] loaded {len(_mb_jpeg_queue)} JPEGs into queue (cleaned orphans)")

        # Add new frame to queue
        _mb_jpeg_queue.append((day_str + "T" + ts_file.replace("-", ":"), fpath))

        # FIFO: if oldest frame is >3hrs old, delete it (one at a time)
        cutoff_str = (dt_ist - _td3(hours=3)).strftime("%Y-%m-%dT%H:%M:%S")
        if _mb_jpeg_queue and _mb_jpeg_queue[0][0] < cutoff_str:
            old_ts, old_path = _mb_jpeg_queue.popleft()
            try:
                if os.path.exists(old_path):
                    os.remove(old_path)
            except Exception:
                pass

        # Always append metadata (tiny, kept forever)
        meta = {"file": fname, "device_id": device_id,
                "camera": camera_name, "heads": head_count, "ts": timestamp_str}
        with open(os.path.join(day_dir, "metadata.jsonl"), "a") as f:
            f.write(json.dumps(meta) + "\n")

        # Append to permanent CSV stats (real camera name, not truncated)
        csv_path = os.path.join(_MB_FRAMES_DIR, "crowd_stats.csv")
        csv_exists = os.path.exists(csv_path)
        csv_cam = camera_name.replace(",", ";")  # escape commas for CSV
        with open(csv_path, "a") as f:
            if not csv_exists:
                f.write("date,hour,device_id,camera,heads,timestamp\n")
            f.write(f"{day_str},{dt_ist.hour},{device_id},{csv_cam},{head_count},{timestamp_str}\n")
    except Exception as e:
        print(f"[magicbox-crowd] frame save error: {e}")



@app.post("/api/magicbox-crowd/ingest")
async def api_magicbox_crowd_ingest(request: Request):
    """Receive head count report from an RK3566 edge device. No auth required."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    device_id = body.get("device_id")
    if not device_id:
        raise HTTPException(status_code=400, detail="missing device_id")
    now = body.get("timestamp", datetime.utcnow().isoformat() + "Z")
    with _mb_crowd_lock:
        if device_id not in _mb_crowd:
            _mb_crowd[device_id] = {"first_seen": now, "total_reports": 0}
        _mb_crowd[device_id]["last_report"] = body
        _mb_crowd[device_id]["last_seen"] = now
        _mb_crowd[device_id]["total_reports"] += 1
        # Recent buffer for API + persist ALL to disk
        for cam in body.get("cameras", []):
            fb = cam.get("frame_b64")
            if fb and cam.get("grab_ok"):
                _mb_crowd_frames.append({
                    "device_id": device_id,
                    "camera_name": cam.get("name", "?"),
                    "head_count": cam.get("head_count", 0),
                    "timestamp": now,
                    "frame_b64": fb,
                })
                _save_frame_to_disk(device_id, cam.get("name", "unknown"),
                                    cam.get("head_count", 0), now, fb)
    return {"status": "ok"}

@app.get("/api/magicbox-crowd/status")
def api_magicbox_crowd_status():
    """Return latest data for all reporting devices. Polled by frontend."""
    import time as _time
    now = _time.time()
    with _mb_crowd_lock:
        devices = []
        for did, entry in _mb_crowd.items():
            report = entry.get("last_report", {})
            # Parse last_seen to check if online
            last_seen_str = entry.get("last_seen", "")
            is_online = False
            try:
                from datetime import datetime as _dt, timezone as _tz
                ls = _dt.fromisoformat(last_seen_str.replace("Z", "+00:00"))
                is_online = (now - ls.timestamp()) < 30
            except Exception:
                pass
            cams = [{k: v for k, v in c.items() if k != "frame_b64"} for c in report.get("cameras", [])]
            total_heads = sum(c.get("head_count", 0) for c in cams if c.get("grab_ok"))
            devices.append({
                "device_id": did,
                "is_online": is_online,
                "last_seen": last_seen_str,
                "total_reports": entry.get("total_reports", 0),
                "total_heads": total_heads,
                "cameras": cams,
                "system": report.get("system", {}),
                "cycle_ms": report.get("cycle_ms", 0),
            })
    return {"devices": devices, "total_devices": len(devices)}

@app.get("/api/magicbox-crowd/frames")
def api_magicbox_crowd_frames(request: Request):
    """Return detection frames (newest first). ?limit=N to cap (default 20)."""
    try:
        limit = int(request.query_params.get("limit", 20))
    except (ValueError, TypeError):
        limit = 20
    limit = max(1, min(limit, 500))
    with _mb_crowd_lock:
        frames = list(reversed(_mb_crowd_frames))[:limit]
    return {"frames": frames, "count": len(frames), "total": len(_mb_crowd_frames)}


# ── MagicBox Crowd: Hourly stats + Historical frames ─────────────────
_hourly_cache: dict = {}  # IST-aware  # {"2026-04-29": {"data": [...], "ts": time.time()}}
_hourly_records_cache: dict = {}  # {"2026-04-30": {"records": [...], "ts": time.time(), "size": N}}

@app.get("/api/magicbox-crowd/hourly")
def api_magicbox_crowd_hourly(request: Request):
    """Hourly head count aggregation from disk metadata. ?date=YYYY-MM-DD&device_id=X&camera=Y"""
    import time as _t
    from datetime import timezone as _tz, timedelta as _td
    _IST = _tz(_td(hours=5, minutes=30))
    date = request.query_params.get("date", datetime.now(_IST).strftime("%Y-%m-%d"))
    device_filter = request.query_params.get("device_id", "")
    device_ids_raw = request.query_params.get("device_ids", "")
    device_ids_set = set(d.strip() for d in device_ids_raw.split(",") if d.strip()) if device_ids_raw else set()
    camera_filter = request.query_params.get("camera", "")
    cache_key = f"{date}:{device_filter}:{device_ids_raw}:{camera_filter}"

    # Check cache (60s for today, indefinite for past dates)
    if cache_key in _hourly_cache:
        cached = _hourly_cache[cache_key]
        is_today = date == datetime.now(_IST).strftime("%Y-%m-%d")
        if not is_today or (_t.time() - cached["ts"]) < 60:
            return cached["data"]

    meta_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "crowd_frames", date, "metadata.jsonl")
    if not os.path.exists(meta_path):
        return {"date": date, "hours": [], "total_frames": 0, "peak_hour": -1}

    # Cache parsed records so filtered queries don't re-read 58MB file
    is_today = date == datetime.now(_IST).strftime("%Y-%m-%d")
    file_size = os.path.getsize(meta_path)
    rc = _hourly_records_cache.get(date)
    if rc and (not is_today or (_t.time() - rc["ts"]) < 30) and rc["size"] == file_size:
        all_records = rc["records"]
    else:
        all_records = []
        try:
            with open(meta_path) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rec = json.loads(line)
                    except Exception:
                        continue
                    ts = rec.get("ts", "")
                    try:
                        _ist_dt = datetime.fromisoformat(ts.replace("Z", "+00:00")).astimezone(_IST)
                        h = _ist_dt.hour
                    except Exception:
                        h = -1
                    all_records.append((rec.get("device_id", ""), rec.get("camera", ""), rec.get("heads", 0), h))
        except Exception as e:
            print(f"[hourly] error reading {meta_path}: {e}")
            return {"date": date, "hours": [], "total_frames": 0, "peak_hour": -1}
        _hourly_records_cache[date] = {"records": all_records, "ts": _t.time(), "size": file_size}

    hours = {h: {"hour": h, "total_heads": 0, "frame_count": 0} for h in range(24)}
    total = 0
    for dev, cam, heads, h in all_records:
        if h < 0:
            continue
        if device_filter and dev != device_filter:
            continue
        if device_ids_set and dev not in device_ids_set:
            continue
        if camera_filter and cam != camera_filter:
            continue
        hours[h]["total_heads"] += heads
        hours[h]["frame_count"] += 1
        total += 1

    hours_list = sorted(hours.values(), key=lambda x: x["hour"])
    peak = max(hours_list, key=lambda x: x["total_heads"])
    result = {"date": date, "hours": hours_list, "total_frames": total, "peak_hour": peak["hour"] if peak["total_heads"] > 0 else -1}
    _hourly_cache[cache_key] = {"data": result, "ts": _t.time()}
    return result


@app.get("/api/magicbox-crowd/frames-history")
def api_magicbox_crowd_frames_history(request: Request):
    """Return historical frames from disk for a specific date+hour. Returns base64 JPEGs."""
    from datetime import timezone as _tz, timedelta as _td
    _IST = _tz(_td(hours=5, minutes=30))
    date = request.query_params.get("date", datetime.now(_IST).strftime("%Y-%m-%d"))
    try:
        hour = int(request.query_params.get("hour", -1))
    except (ValueError, TypeError):
        hour = -1
    device_filter = request.query_params.get("device_id", "")
    device_ids_raw = request.query_params.get("device_ids", "")
    device_ids_set = set(d.strip() for d in device_ids_raw.split(",") if d.strip()) if device_ids_raw else set()
    camera_filter = request.query_params.get("camera", "")
    try:
        limit = int(request.query_params.get("limit", 30))
    except (ValueError, TypeError):
        limit = 30
    limit = max(1, min(limit, 1000))
    try:
        min_heads = int(request.query_params.get("min_heads", 0))
    except (ValueError, TypeError):
        min_heads = 0

    day_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "crowd_frames", date)
    meta_path = os.path.join(day_dir, "metadata.jsonl")
    if not os.path.exists(meta_path):
        return {"frames": [], "count": 0}

    import base64 as _b64
    matching = []
    try:
        with open(meta_path) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                rec_dev = rec.get("device_id", "")
                if device_filter and rec_dev != device_filter:
                    continue
                if device_ids_set and rec_dev not in device_ids_set:
                    continue
                if camera_filter and rec.get("camera") != camera_filter:
                    continue
                ts = rec.get("ts", "")
                if hour >= 0:
                    try:
                        _ist_dt = datetime.fromisoformat(ts.replace("Z", "+00:00")).astimezone(_IST)
                        rec_hour = _ist_dt.hour
                    except Exception:
                        continue
                    if rec_hour != hour:
                        continue
                if min_heads > 0 and rec.get("heads", 0) < min_heads:
                    continue
                if min_heads > 0 and rec.get("heads", 0) < min_heads:
                    continue
                matching.append(rec)
    except Exception as e:
        print(f"[frames-history] error: {e}")
        return {"frames": [], "count": 0}

    # Take last N (newest)
    matching = matching[-limit:]
    matching.reverse()

    MAX_JPEG_SIZE = 500 * 1024  # skip files > 500KB (likely corrupted)

    def _read_frame(rec):
        fp = os.path.join(day_dir, rec.get("file", ""))
        b64 = ""
        if rec.get("file") and os.path.exists(fp):
            try:
                sz = os.path.getsize(fp)
                if 0 < sz <= MAX_JPEG_SIZE:
                    with open(fp, "rb") as f:
                        b64 = _b64.b64encode(f.read()).decode("ascii")
            except Exception:
                pass
        return {
            "device_id": rec.get("device_id", ""),
            "camera_name": rec.get("camera", ""),
            "head_count": rec.get("heads", 0),
            "timestamp": rec.get("ts", ""),
            "frame_b64": b64,
        }

    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=8) as pool:
        frames = list(pool.map(_read_frame, matching))

    return {"frames": frames, "count": len(frames), "total_matching": len(matching)}


@app.get("/api/magicbox-crowd/fleet")
async def api_magicbox_crowd_fleet():
    """Return static fleet data from JSON file (from magicbox_devices.xlsx)."""
    global _mb_fleet_cache
    if _mb_fleet_cache["data"] is not None:
        return {"stations": _mb_fleet_cache["data"]}
    try:
        fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "magicbox_fleet.json")
        with open(fpath) as f:
            _mb_fleet_cache["data"] = json.load(f)
        return {"stations": _mb_fleet_cache["data"]}
    except Exception as e:
        print(f"[fleet] error loading fleet JSON: {e}")
        return {"stations": []}
